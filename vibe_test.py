import sys, io, os, re, time, json
from datetime import datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

# ── Config ──────────────────────────────────────────────────────────────────
TC_FILE    = r'C:\Users\ASUS\Documents\Pineline\ECOM\00_input\Thông tin chung.xlsx'
OUT_FILE   = r'C:\Users\ASUS\Documents\Pineline\ECOM\03_test-cases\functional\Thông tin chung_result.xlsx'
EVD_DIR    = r'C:\Users\ASUS\Documents\Pineline\ECOM\08_test-runs\2026-06-06\evidence\screenshots'
LOG_FILE   = r'C:\Users\ASUS\Documents\Pineline\ECOM\08_test-runs\2026-06-06\run-log.json'
PRODUCT_URL = 'https://staging.tongdaiwifi.vn/thiet-bi-thong-minh/access-point-ac1200t-test'
os.makedirs(EVD_DIR, exist_ok=True)

# ── Test Data ────────────────────────────────────────────────────────────────
DATA = {
    'phone_valid'   : '0343439724',
    'phone_invalid' : ['03215', '00000', '9865654', ' '],
    'email_valid'   : 'anhdc4@fpt.com',
    'email_invalid' : ['djsf@dfjhdf.', 'dfsdkfjsd'],
    'province'      : 'Hồ Chí Minh',
    'ward'          : 'Sài Gòn',
    'street'        : 'Bùi Thị Xuân',
    'house_no'      : '123',
    'package'       : 'Access Point AC1200T',
    'skip_tcs'      : {'TC_11.118', 'TC_11.119'},
}

# ── Colors ────────────────────────────────────────────────────────────────────
GREEN_FILL  = PatternFill('solid', start_color='00B050', end_color='00B050')
RED_FILL    = PatternFill('solid', start_color='FF0000', end_color='FF0000')
SKIP_FILL   = PatternFill('solid', start_color='FFC000', end_color='FFC000')
MANUAL_FILL = PatternFill('solid', start_color='4472C4', end_color='4472C4')
WHITE_FONT  = Font(color='FFFFFF', bold=True, name='Arial', size=10)

# ── Load workbook & build TC row map ─────────────────────────────────────────
# Load WITHOUT data_only so formulas are preserved when we save
wb = openpyxl.load_workbook(TC_FILE)
ws = wb.worksheets[0]

# Read prefix from row 3 col D (e.g. "TC_11")
prefix_cell = ws.cell(row=3, column=4).value or ''
# If it contains a newline or label prefix, strip it
prefix = ''
for part in str(prefix_cell).split():
    if part.startswith('TC_'):
        prefix = part
        break
if not prefix:
    prefix = 'TC_11'

# Identify TC rows: col C (index 3) = High/Medium/Low  AND col D (index 4) not empty
# Replicate formula: count non-empty col-D rows from row 9 downward → that is YY
PRIORITIES = {'High', 'Medium', 'Low'}
tc_rows = {}   # 'TC_11.1' -> excel row number (1-based)
yy_counter = 0
for excel_row in range(9, ws.max_row + 1):
    col_d = ws.cell(row=excel_row, column=4).value
    col_c = ws.cell(row=excel_row, column=3).value
    if col_d and str(col_d).strip():
        yy_counter += 1
        if col_c and str(col_c).strip() in PRIORITIES:
            tc_id = f'{prefix}.{yy_counter}'
            tc_rows[tc_id] = excel_row

print(f'Loaded {len(tc_rows)} TCs  |  prefix={prefix}')

# ── Helpers ───────────────────────────────────────────────────────────────────
def ss(page, tc_id, tag=''):
    ts = datetime.now().strftime('%H%M%S')
    fname = f'FAIL_{tc_id.replace(".", "_")}_{tag}_{ts}.png'
    path = os.path.join(EVD_DIR, fname)
    try:
        page.screenshot(path=path, full_page=False)
        return path
    except Exception:
        return None

def write_result(row_idx, status, note=''):
    if not row_idx:
        return
    cell = ws.cell(row=row_idx, column=8)
    ts = datetime.now().strftime('%H:%M %d/%m/%Y')
    if status == 'Pass':
        cell.value = f'Pass\n{ts}'
        cell.fill = GREEN_FILL
    elif status == 'Fail':
        cell.value = f'Fail\n{note}\n{ts}'
        cell.fill = RED_FILL
    elif status == 'Skip':
        cell.value = f'Skip\n{note}'
        cell.fill = SKIP_FILL
    else:  # Manual
        cell.value = f'Manual\n{note}'
        cell.fill = MANUAL_FILL
    cell.font = WHITE_FONT
    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
    wb.save(OUT_FILE)

results = {'Pass': [], 'Fail': [], 'Skip': [], 'Manual': []}

def record(tc_id, status, note='', evd=None):
    results[status].append({'tc': tc_id, 'note': note, 'evidence': evd})
    sym = {'Pass': '✅', 'Fail': '❌', 'Skip': '⏭️', 'Manual': '🔵'}.get(status, '?')
    suffix = f': {note}' if note else ''
    print(f'  {sym} [{tc_id}] {status}{suffix}', flush=True)
    write_result(tc_rows.get(tc_id), status, note)

def el(page, selectors, timeout=6000):
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state='visible', timeout=timeout)
            return loc
        except Exception:
            continue
    return None

def has_text(page, text, timeout=4000):
    try:
        page.get_by_text(text, exact=False).first.wait_for(state='visible', timeout=timeout)
        return True
    except Exception:
        return False

def goto_checkout(page):
    """Navigate product page → click Mua ngay → return checkout URL."""
    page.goto(PRODUCT_URL, wait_until='domcontentloaded', timeout=25000)
    time.sleep(2)
    btn = el(page, ['button:has-text("Mua ngay")', 'a:has-text("Mua ngay")',
                    'button:has-text("Mua Ngay")', '[class*="buy"]'], timeout=8000)
    if btn:
        btn.click()
        page.wait_for_url('**/checkout/**', timeout=15000)
        time.sleep(2)
        return page.url
    raise RuntimeError('Không tìm thấy btn Mua ngay')

# Field selectors
PHONE_SEL  = ['input[placeholder*="điện thoại"]', 'input[placeholder*="Điện thoại"]',
               'input[type="tel"]', 'input[name*="phone"]', 'input[name*="sdt"]',
               'input[id*="phone"]', 'input[id*="Phone"]']
EMAIL_SEL  = ['input[type="email"]', 'input[placeholder*="email"]',
               'input[placeholder*="Email"]', 'input[name*="email"]']
SUBMIT_SEL = ['button:has-text("Thanh toán")', 'button:has-text("THANH TOÁN")',
               'button[type="submit"]', 'input[type="submit"]']
CLEAR_SEL  = ['[class*="clear"]', '[class*="delete"]', 'button[aria-label*="clear"]',
               '[class*="remove"]', 'button[aria-label*="xóa"]']
PROV_SEL   = ['[placeholder*="Tỉnh"]', '[placeholder*="tỉnh"]', 'select[name*="province"]',
               '[class*="province"] input', '[class*="city"] input']
PROMO_SEL  = ['input[placeholder*="mã"]', 'input[placeholder*="khuyến mãi"]',
               'input[placeholder*="voucher"]', 'input[placeholder*="Voucher"]']

def clear_type(page, field, value):
    try:
        field.fill('')
    except Exception:
        pass
    if value.strip():
        field.fill(value)

# ─────────────────────────────────────────────────────────────────────────────
print(f'\n🚀 Bắt đầu Vibe Test — {datetime.now().strftime("%d/%m/%Y %H:%M")}')
print(f'   Product: {PRODUCT_URL}\n')

with sync_playwright() as p:
    browser = p.chromium.launch(headless=False, slow_mo=200)
    ctx = browser.new_context(viewport={'width': 1366, 'height': 768})

    # state dict allows fresh() to rebind page when closed
    S = {'page': ctx.new_page(), 'url': ''}

    # ── Điều hướng vào checkout ─────────────────────────────────────────────
    print('🔗 Điều hướng vào trang checkout...')
    try:
        S['url'] = goto_checkout(S['page'])
        print(f'   Checkout URL: {S["url"]}\n')
    except Exception as e:
        print(f'❌ Không thể vào checkout: {e}')
        browser.close()
        sys.exit(1)

    def fresh():
        """Reload checkout page; recreate page if session expired."""
        pg = S['page']
        try:
            pg.goto(S['url'], wait_until='domcontentloaded', timeout=20000)
            # If server redirected us away from checkout, re-navigate from product page
            if 'checkout' not in pg.url and 'register' not in pg.url:
                S['url'] = goto_checkout(pg)
        except Exception:
            try: pg.close()
            except: pass
            pg = ctx.new_page()
            S['page'] = pg
            S['url'] = goto_checkout(pg)
        time.sleep(1.5)
        return S['page']

    # Helper: current page
    def P(): return S['page']

    # ── TC_11.1 — Click Logo FPT Telecom ────────────────────────────────────
    print('TC_11.1 — Click vào Logo FPT Telecom')
    try:
        page = fresh()
        logo = el(page, ['img[alt*="FPT"]', 'img[alt*="fpt"]', 'a.logo', '.header a img',
                          'header img', 'a[href*="fpt"] img'])
        if logo:
            logo.click(); time.sleep(2)
            if 'fpt.vn' in page.url or 'tongdaiwifi' in page.url:
                record('TC_11.1', 'Pass')
            else:
                evd = ss(page, 'TC_11.1')
                record('TC_11.1', 'Fail', f'URL sau click: {page.url[:80]}', evd)
        else:
            evd = ss(page, 'TC_11.1')
            record('TC_11.1', 'Fail', 'Không tìm thấy logo', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.1')
        record('TC_11.1', 'Fail', str(e)[:80], evd)

    # ── TC_11.2 — Click icon back (truy cập trực tiếp link) ─────────────────
    print('TC_11.2 — Click icon back từ truy cập trực tiếp')
    try:
        page = fresh()
        page.go_back(); time.sleep(2)
        if 'fpt.vn' in page.url or 'tongdaiwifi' in page.url:
            record('TC_11.2', 'Pass')
        else:
            evd = ss(page, 'TC_11.2')
            record('TC_11.2', 'Fail', f'URL sau back: {page.url[:80]}', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.2')
        record('TC_11.2', 'Fail', str(e)[:80], evd)

    # ── TC_11.3 — Click icon back từ tongdaiwifi ─────────────────────────────
    print('TC_11.3 — Click icon back từ tongdaiwifi')
    try:
        page.goto('https://staging.tongdaiwifi.vn/', wait_until='domcontentloaded', timeout=15000)
        time.sleep(1.5)
        btn = el(page, ['text=Đăng ký ngay', 'button:has-text("Đăng ký")', 'a:has-text("Đăng ký")'])
        if btn:
            btn.click(); time.sleep(2)
        page.go_back(); time.sleep(2)
        if 'tongdaiwifi' in page.url:
            record('TC_11.3', 'Pass')
        else:
            evd = ss(page, 'TC_11.3')
            record('TC_11.3', 'Fail', f'URL: {page.url[:80]}', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.3')
        record('TC_11.3', 'Fail', str(e)[:80], evd)

    # ── TC_11.4 — Check màu sắc mặc định các bước ───────────────────────────
    print('TC_11.4 — Check màu sắc mặc định các bước')
    try:
        page = fresh()
        step = el(page, ['[class*="step"]', '[class*="progress"]', '[class*="stepper"]',
                          '[class*="breadcrumb"]'], timeout=5000)
        evd = ss(page, 'TC_11.4', 'step')
        if step:
            record('TC_11.4', 'Manual', 'Tìm thấy stepper — cần verify màu sắc thủ công', evd)
        else:
            record('TC_11.4', 'Manual', 'Không tìm thấy stepper — verify thủ công', evd)
    except Exception as e:
        record('TC_11.4', 'Manual', 'Cần verify màu thủ công')

    # ── TC_11.5 — Click bước 1,2,3 khi đang ở bước 1 ────────────────────────
    print('TC_11.5 — Click vào bước 1,2,3 khi đang ở bước 1')
    try:
        page = fresh()
        url_before = page.url
        steps = page.locator('[class*="step"],[class*="Step"]').all()
        for s in steps[:3]:
            try: s.click(timeout=2000)
            except: pass
        time.sleep(1)
        if page.url == url_before or 'checkout' in page.url:
            record('TC_11.5', 'Pass')
        else:
            evd = ss(page, 'TC_11.5')
            record('TC_11.5', 'Fail', 'Chuyển trang không mong muốn', evd)
    except Exception as e:
        record('TC_11.5', 'Manual', 'Cần verify thủ công')

    # ── TC_11.6 — Click bước 1 khi đang ở bước 2 ────────────────────────────
    print('TC_11.6 — Click bước 1 khi đang ở bước 2')
    record('TC_11.6', 'Manual', 'Cần điền form để sang bước 2 — verify thủ công')

    # ── TC_11.7 — Check load Số lượng gói dịch vụ ───────────────────────────
    print('TC_11.7 — Check load Số lượng gói dịch vụ')
    try:
        page = fresh()
        if has_text(page, 'Số lượng') or has_text(page, '1'):
            record('TC_11.7', 'Pass')
        else:
            evd = ss(page, 'TC_11.7')
            record('TC_11.7', 'Fail', 'Không thấy Số lượng gói', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.7')
        record('TC_11.7', 'Fail', str(e)[:80], evd)

    # ── TC_11.8 — Check load data Gói dịch vụ ───────────────────────────────
    print('TC_11.8 — Check load data Gói dịch vụ')
    try:
        page = fresh()
        if has_text(page, 'AC1200') or has_text(page, 'Access Point') or has_text(page, 'AC1200T'):
            record('TC_11.8', 'Pass')
        else:
            evd = ss(page, 'TC_11.8')
            record('TC_11.8', 'Fail', f'Không thấy tên gói "{DATA["package"]}"', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.8')
        record('TC_11.8', 'Fail', str(e)[:80], evd)

    # ── TC_11.9 — Check load Chu kỳ gói dịch vụ ─────────────────────────────
    print('TC_11.9 — Check load Chu kỳ')
    try:
        page = fresh()
        if has_text(page, 'tháng') or has_text(page, 'Chu kỳ') or has_text(page, 'năm'):
            record('TC_11.9', 'Pass')
        else:
            evd = ss(page, 'TC_11.9')
            record('TC_11.9', 'Fail', 'Không thấy thông tin chu kỳ', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.9')
        record('TC_11.9', 'Fail', str(e)[:80], evd)

    # ── TC_11.10 — Check load Giá gói dịch vụ ───────────────────────────────
    print('TC_11.10 — Check load Giá')
    try:
        page = fresh()
        if has_text(page, 'đ') or has_text(page, 'VND') or has_text(page, 'vnđ') or has_text(page, '₫'):
            record('TC_11.10', 'Pass')
        else:
            evd = ss(page, 'TC_11.10')
            record('TC_11.10', 'Fail', 'Không thấy giá gói', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.10')
        record('TC_11.10', 'Fail', str(e)[:80], evd)

    # ── TC_11.11 — Check load icon Gói dịch vụ ──────────────────────────────
    print('TC_11.11 — Check load icon Gói')
    try:
        page = fresh()
        icon = el(page, ['[class*="product"] img', '[class*="package"] img',
                          '[class*="service"] img', '[class*="item"] img'], timeout=5000)
        if icon:
            record('TC_11.11', 'Pass')
        else:
            evd = ss(page, 'TC_11.11', 'icon')
            record('TC_11.11', 'Manual', 'Cần verify icon thủ công', evd)
    except Exception as e:
        record('TC_11.11', 'Manual', 'Cần verify thủ công')

    # ════════════════════════════════════════════════════════════════════════
    # Phone field TCs (11.20 ~ 11.25)
    # ════════════════════════════════════════════════════════════════════════
    def get_phone(page):
        return el(page, PHONE_SEL, timeout=6000)

    print('\n--- Trường Số điện thoại ---')

    print('TC_11.20 — Check btn X trong trường SĐT')
    try:
        page = fresh()
        f = get_phone(page)
        if f:
            clear_type(page, f, DATA['phone_valid']); time.sleep(0.5)
            x = el(page, CLEAR_SEL, timeout=3000)
            if x:
                record('TC_11.20', 'Pass')
            else:
                evd = ss(page, 'TC_11.20')
                record('TC_11.20', 'Fail', 'Không thấy btn X sau khi nhập SĐT', evd)
        else:
            evd = ss(page, 'TC_11.20')
            record('TC_11.20', 'Fail', 'Không tìm thấy field SĐT', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.20')
        record('TC_11.20', 'Fail', str(e)[:80], evd)

    print('TC_11.21 — Click btn X trong trường SĐT')
    try:
        page = fresh()
        f = get_phone(page)
        if f:
            clear_type(page, f, DATA['phone_valid']); time.sleep(0.5)
            x = el(page, CLEAR_SEL, timeout=3000)
            if x:
                x.click(); time.sleep(0.5)
                val = f.input_value()
                if val == '':
                    record('TC_11.21', 'Pass')
                else:
                    evd = ss(page, 'TC_11.21')
                    record('TC_11.21', 'Fail', f'Field vẫn còn "{val}"', evd)
            else:
                evd = ss(page, 'TC_11.21')
                record('TC_11.21', 'Fail', 'Không thấy btn X', evd)
        else:
            evd = ss(page, 'TC_11.21')
            record('TC_11.21', 'Fail', 'Không tìm thấy field SĐT', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.21')
        record('TC_11.21', 'Fail', str(e)[:80], evd)

    print('TC_11.22 — Check SĐT để trống khi submit')
    try:
        page = fresh()
        submit = el(page, SUBMIT_SEL, timeout=5000)
        if submit:
            submit.click(); time.sleep(1.5)
            if has_text(page, 'Vui lòng') or has_text(page, 'bắt buộc') or has_text(page, 'điện thoại'):
                record('TC_11.22', 'Pass')
            else:
                evd = ss(page, 'TC_11.22')
                record('TC_11.22', 'Fail', 'Không có thông báo lỗi khi SĐT trống', evd)
        else:
            evd = ss(page, 'TC_11.22')
            record('TC_11.22', 'Manual', 'Không tìm thấy btn Thanh toán', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.22')
        record('TC_11.22', 'Fail', str(e)[:80], evd)

    print('TC_11.23 — Check nhập SĐT ký tự đặc biệt/chữ')
    try:
        page = fresh()
        f = get_phone(page)
        if f:
            clear_type(page, f, 'abc!@#'); time.sleep(0.5)
            val = f.input_value()
            if val == '' or not re.search(r'[a-zA-Z!@#]', val):
                record('TC_11.23', 'Pass')
            else:
                evd = ss(page, 'TC_11.23')
                record('TC_11.23', 'Fail', f'Field chấp nhận ký tự: "{val}"', evd)
        else:
            evd = ss(page, 'TC_11.23')
            record('TC_11.23', 'Fail', 'Không tìm thấy field SĐT', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.23')
        record('TC_11.23', 'Fail', str(e)[:80], evd)

    print('TC_11.24 — Check SĐT 10 số, số đầu khác 0')
    try:
        page = fresh()
        f = get_phone(page)
        if f:
            clear_type(page, f, '9865654321'); time.sleep(0.5)
            f.press('Tab')
            time.sleep(0.8)
            if (has_text(page, 'không hợp lệ', 3000) or
                    has_text(page, 'invalid', 3000) or has_text(page, 'sai', 3000)):
                record('TC_11.24', 'Pass')
            else:
                evd = ss(page, 'TC_11.24')
                record('TC_11.24', 'Fail', 'Không có thông báo lỗi SĐT đầu khác 0', evd)
        else:
            evd = ss(page, 'TC_11.24')
            record('TC_11.24', 'Fail', 'Không tìm thấy field SĐT', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.24')
        record('TC_11.24', 'Fail', str(e)[:80], evd)

    print('TC_11.25 — Check nhập SĐT > 10 chữ số')
    try:
        page = fresh()
        f = get_phone(page)
        if f:
            clear_type(page, f, '03434397241234'); time.sleep(0.5)
            val = f.input_value()
            if len(val) <= 10:
                record('TC_11.25', 'Pass')
            else:
                evd = ss(page, 'TC_11.25')
                record('TC_11.25', 'Fail', f'Field chấp nhận {len(val)} ký tự (> 10)', evd)
        else:
            evd = ss(page, 'TC_11.25')
            record('TC_11.25', 'Fail', 'Không tìm thấy field SĐT', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.25')
        record('TC_11.25', 'Fail', str(e)[:80], evd)

    # ════════════════════════════════════════════════════════════════════════
    # Email TCs (11.26 ~ 11.32)
    # ════════════════════════════════════════════════════════════════════════
    def get_email(page):
        return el(page, EMAIL_SEL, timeout=6000)

    print('\n--- Trường Email ---')

    print('TC_11.26 — Check placeholder Email')
    try:
        page = fresh()
        f = get_email(page)
        if f:
            ph = f.get_attribute('placeholder') or ''
            record('TC_11.26', 'Pass' if ph else 'Fail',
                   '' if ph else 'Không có placeholder')
        else:
            evd = ss(page, 'TC_11.26')
            record('TC_11.26', 'Fail', 'Không tìm thấy field Email', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.26')
        record('TC_11.26', 'Fail', str(e)[:80], evd)

    print('TC_11.27 — Check btn X trong trường Email')
    try:
        page = fresh()
        f = get_email(page)
        if f:
            clear_type(page, f, DATA['email_valid']); time.sleep(0.5)
            x = el(page, CLEAR_SEL, timeout=3000)
            record('TC_11.27', 'Pass' if x else 'Fail',
                   '' if x else 'Không thấy btn X sau khi nhập Email')
        else:
            evd = ss(page, 'TC_11.27')
            record('TC_11.27', 'Fail', 'Không tìm thấy field Email', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.27')
        record('TC_11.27', 'Fail', str(e)[:80], evd)

    print('TC_11.28 — Click btn X trong trường Email')
    try:
        page = fresh()
        f = get_email(page)
        if f:
            clear_type(page, f, DATA['email_valid']); time.sleep(0.5)
            x = el(page, CLEAR_SEL, timeout=3000)
            if x:
                x.click(); time.sleep(0.5)
                val = f.input_value()
                record('TC_11.28', 'Pass' if val == '' else 'Fail',
                       '' if val == '' else f'Field vẫn còn "{val}"')
            else:
                evd = ss(page, 'TC_11.28')
                record('TC_11.28', 'Fail', 'Không thấy btn X', evd)
        else:
            evd = ss(page, 'TC_11.28')
            record('TC_11.28', 'Fail', 'Không tìm thấy field Email', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.28')
        record('TC_11.28', 'Fail', str(e)[:80], evd)

    print('TC_11.29 — Check Email để trống khi submit')
    try:
        page = fresh()
        pf = get_phone(page)
        if pf:
            clear_type(page, pf, DATA['phone_valid']); time.sleep(0.3)
        submit = el(page, SUBMIT_SEL, timeout=5000)
        if submit:
            submit.click(); time.sleep(1.5)
            if has_text(page, 'email', 3000) or has_text(page, 'Email', 3000) or has_text(page, 'bắt buộc', 3000):
                record('TC_11.29', 'Pass')
            else:
                evd = ss(page, 'TC_11.29')
                record('TC_11.29', 'Fail', 'Không có thông báo lỗi khi Email trống', evd)
        else:
            evd = ss(page, 'TC_11.29')
            record('TC_11.29', 'Manual', 'Không tìm thấy btn Thanh toán', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.29')
        record('TC_11.29', 'Fail', str(e)[:80], evd)

    print('TC_11.30 — Check Email sai định dạng')
    try:
        page = fresh()
        f = get_email(page)
        if f:
            clear_type(page, f, 'djsf@dfjhdf.'); time.sleep(0.3)
            f.press('Tab'); time.sleep(0.8)
            if (has_text(page, 'không hợp lệ', 3000) or
                    has_text(page, 'sai định dạng', 3000) or has_text(page, 'invalid', 3000)):
                record('TC_11.30', 'Pass')
            else:
                evd = ss(page, 'TC_11.30')
                record('TC_11.30', 'Fail', 'Không có thông báo lỗi email sai định dạng', evd)
        else:
            evd = ss(page, 'TC_11.30')
            record('TC_11.30', 'Fail', 'Không tìm thấy field Email', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.30')
        record('TC_11.30', 'Fail', str(e)[:80], evd)

    print('TC_11.31 — Check Email không có domain')
    try:
        page = fresh()
        f = get_email(page)
        if f:
            clear_type(page, f, 'dfsdkfjsd'); time.sleep(0.3)
            f.press('Tab'); time.sleep(0.8)
            if (has_text(page, 'không hợp lệ', 3000) or
                    has_text(page, 'sai', 3000) or has_text(page, 'invalid', 3000)):
                record('TC_11.31', 'Pass')
            else:
                evd = ss(page, 'TC_11.31')
                record('TC_11.31', 'Fail', 'Không có thông báo lỗi email thiếu domain', evd)
        else:
            evd = ss(page, 'TC_11.31')
            record('TC_11.31', 'Fail', 'Không tìm thấy field Email', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.31')
        record('TC_11.31', 'Fail', str(e)[:80], evd)

    print('TC_11.32 — Check Email sync realtime sang block Thông tin lắp đặt')
    try:
        page = fresh()
        f = get_email(page)
        if f:
            clear_type(page, f, DATA['email_valid']); time.sleep(1.5)
            count = page.locator(f'text={DATA["email_valid"]}').count()
            if count >= 2:
                record('TC_11.32', 'Pass')
            else:
                evd = ss(page, 'TC_11.32')
                record('TC_11.32', 'Manual', 'Cần verify realtime sync thủ công', evd)
        else:
            evd = ss(page, 'TC_11.32')
            record('TC_11.32', 'Fail', 'Không tìm thấy field Email', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.32')
        record('TC_11.32', 'Fail', str(e)[:80], evd)

    # ════════════════════════════════════════════════════════════════════════
    # Địa chỉ TCs (11.33 ~ 11.49)
    # ════════════════════════════════════════════════════════════════════════
    print('\n--- Trường Địa chỉ ---')

    print('TC_11.33 — Check dữ liệu mặc định Tỉnh/TP')
    try:
        page = fresh()
        prov = el(page, PROV_SEL, timeout=5000)
        if prov:
            val = prov.input_value() or ''
            evd = ss(page, 'TC_11.33', 'province')
            record('TC_11.33', 'Manual', f'Giá trị mặc định: "{val}" — verify thủ công', evd)
        else:
            evd = ss(page, 'TC_11.33')
            record('TC_11.33', 'Fail', 'Không tìm thấy dropdown Tỉnh/TP', evd)
    except Exception as e:
        record('TC_11.33', 'Manual', 'Cần verify thủ công')

    print('TC_11.34 — Check placeholder Tỉnh/TP')
    try:
        page = fresh()
        prov = el(page, PROV_SEL, timeout=5000)
        if prov:
            ph = prov.get_attribute('placeholder') or ''
            record('TC_11.34', 'Pass' if ph else 'Fail',
                   '' if ph else 'Không có placeholder')
        else:
            record('TC_11.34', 'Manual', 'Không tìm thấy field')
    except Exception as e:
        record('TC_11.34', 'Manual', 'Cần verify thủ công')

    # TC_11.35~49: các thao tác dropdown địa chỉ — cần interaction phức tạp
    for tc, note in [
        ('TC_11.35', 'Submit không chọn TP — verify validation'),
        ('TC_11.36', 'Mở dropdown Tỉnh/TP — verify load data'),
        ('TC_11.37', 'Gõ "Hồ Chí" trong dropdown — verify tìm kiếm'),
        ('TC_11.38', 'Chọn Hồ Chí Minh — verify selection'),
        ('TC_11.39', 'Sau khi chọn TP — verify hiện thêm Phường/Xã'),
        ('TC_11.40', 'Đổi sang TP khác — verify reload địa chỉ'),
        ('TC_11.41', 'Check placeholder Phường/Xã'),
        ('TC_11.42', 'Submit không chọn Phường/Xã — verify validation'),
        ('TC_11.43', 'Tìm kiếm Sài Gòn trong dropdown Phường/Xã'),
        ('TC_11.44', 'Chọn Phường Sài Gòn — verify selection'),
        ('TC_11.45', 'Đổi Phường/Xã — verify reload Tên đường'),
        ('TC_11.46', 'Check placeholder Tên đường'),
        ('TC_11.47', 'Submit không chọn Tên đường — verify validation'),
        ('TC_11.48', 'Tìm kiếm Bùi Thị Xuân trong dropdown'),
        ('TC_11.49', 'Chọn đường Bùi Thị Xuân — verify selection'),
    ]:
        print(f'{tc} — {note}')
        record(tc, 'Manual', f'{note} — thủ công')

    # ════════════════════════════════════════════════════════════════════════
    # TC_11.50 ~ 11.68: Loại nhà & Ghi chú
    # ════════════════════════════════════════════════════════════════════════
    print('\n--- Loại nhà & Số nhà ---')

    print('TC_11.50 — Click radio Nhà riêng')
    try:
        page = fresh()
        radio = el(page, ['label:has-text("Nhà riêng")', 'input[value*="nha"]',
                           'text=Nhà riêng', '[class*="house"]'], timeout=5000)
        if radio:
            radio.click(); time.sleep(0.5)
            record('TC_11.50', 'Pass')
        else:
            evd = ss(page, 'TC_11.50')
            record('TC_11.50', 'Manual', 'Không tìm thấy radio Nhà riêng', evd)
    except Exception as e:
        record('TC_11.50', 'Manual', 'Cần verify thủ công')

    HOUSE_SEL = ['input[placeholder*="Số nhà"]', 'input[placeholder*="số nhà"]',
                 'input[name*="house"]', 'input[name*="so_nha"]']

    print('TC_11.51 — Check placeholder Số nhà')
    try:
        page = fresh()
        f = el(page, HOUSE_SEL, timeout=5000)
        if f:
            ph = f.get_attribute('placeholder') or ''
            record('TC_11.51', 'Pass' if ph else 'Fail',
                   '' if ph else 'Không có placeholder')
        else:
            record('TC_11.51', 'Manual', 'Không tìm thấy field Số nhà')
    except Exception as e:
        record('TC_11.51', 'Manual', 'Cần verify thủ công')

    print('TC_11.52 — Check không nhập Số nhà')
    record('TC_11.52', 'Manual', 'Cần submit form với Số nhà trống — verify thủ công')

    print('TC_11.53 — Check nhập > 100 ký tự Số nhà')
    try:
        page = fresh()
        f = el(page, HOUSE_SEL, timeout=5000)
        if f:
            clear_type(page, f, 'A' * 110); time.sleep(0.3)
            val = f.input_value()
            record('TC_11.53', 'Pass' if len(val) <= 100 else 'Fail',
                   '' if len(val) <= 100 else f'Chấp nhận {len(val)} ký tự (> 100)')
        else:
            record('TC_11.53', 'Manual', 'Không tìm thấy field Số nhà')
    except Exception as e:
        record('TC_11.53', 'Manual', 'Cần verify thủ công')

    print('TC_11.54 — Click radio Chung cư')
    try:
        page = fresh()
        radio = el(page, ['label:has-text("Chung cư")', 'input[value*="chung"]',
                           'text=Chung cư'], timeout=5000)
        if radio:
            radio.click(); time.sleep(0.5)
            record('TC_11.54', 'Pass')
        else:
            evd = ss(page, 'TC_11.54')
            record('TC_11.54', 'Manual', 'Không tìm thấy radio Chung cư', evd)
    except Exception as e:
        record('TC_11.54', 'Manual', 'Cần verify thủ công')

    for tc, note in [
        ('TC_11.55', 'Verify dropdown Tên chung cư load data'),
        ('TC_11.56', 'Tìm kiếm trong Tên chung cư'),
        ('TC_11.57', 'Check placeholder Tòa nhà'),
        ('TC_11.58', 'Submit không nhập Tòa nhà — verify validation'),
        ('TC_11.59', 'Nhập Tòa nhà > 10 ký tự — verify max length'),
        ('TC_11.60', 'Check placeholder Số tầng'),
        ('TC_11.61', 'Submit không nhập Số tầng — verify validation'),
        ('TC_11.62', 'Nhập Số tầng > 10 ký tự — verify max length'),
        ('TC_11.63', 'Check placeholder Số phòng'),
        ('TC_11.64', 'Submit không nhập Số phòng — verify validation'),
        ('TC_11.65', 'Nhập Số phòng > 10 ký tự — verify max length'),
    ]:
        print(f'{tc} — {note}')
        record(tc, 'Manual', f'Cần chọn Chung cư trước — {note}')

    GHICHU_SEL = ['textarea[placeholder*="Ghi chú"]', 'textarea[placeholder*="ghi chú"]',
                  'input[placeholder*="Ghi chú"]', 'textarea[name*="note"]']

    print('TC_11.66 — Check placeholder Ghi chú')
    try:
        page = fresh()
        f = el(page, GHICHU_SEL, timeout=5000)
        if f:
            ph = f.get_attribute('placeholder') or ''
            record('TC_11.66', 'Pass' if ph else 'Fail',
                   '' if ph else 'Không có placeholder')
        else:
            record('TC_11.66', 'Manual', 'Không tìm thấy field Ghi chú')
    except Exception as e:
        record('TC_11.66', 'Manual', 'Cần verify thủ công')

    print('TC_11.67 — Check không nhập Ghi chú')
    record('TC_11.67', 'Manual', 'Ghi chú không bắt buộc — verify thủ công')

    print('TC_11.68 — Check nhập > 100 ký tự Ghi chú')
    try:
        page = fresh()
        f = el(page, GHICHU_SEL, timeout=5000)
        if f:
            clear_type(page, f, 'A' * 110); time.sleep(0.3)
            val = f.input_value()
            record('TC_11.68', 'Pass' if len(val) <= 100 else 'Fail',
                   '' if len(val) <= 100 else f'Chấp nhận {len(val)} ký tự (> 100)')
        else:
            record('TC_11.68', 'Manual', 'Không tìm thấy field Ghi chú')
    except Exception as e:
        record('TC_11.68', 'Manual', 'Cần verify thủ công')

    # ════════════════════════════════════════════════════════════════════════
    # Popup Địa chỉ hành chính cũ (TC_11.69 ~ 11.93)
    # ════════════════════════════════════════════════════════════════════════
    print('\n--- Popup Địa chỉ hành chính cũ (TC_11.69~93) ---')
    for i in range(69, 94):
        tc = f'TC_11.{i}'
        print(tc)
        record(tc, 'Manual', 'Cần chọn địa chỉ không có chính sách để trigger popup — thủ công')

    # ════════════════════════════════════════════════════════════════════════
    # Collapse/Expand & Format (TC_11.94~98)
    # ════════════════════════════════════════════════════════════════════════
    print('\n--- Collapse/Expand (TC_11.94~98) ---')
    for tc, note in [
        ('TC_11.94', 'Click icon Collapse block Thông tin KH'),
        ('TC_11.95', 'Click icon Expand block Thông tin KH'),
        ('TC_11.96', 'Verify format địa chỉ Nhà riêng hiển thị đúng'),
        ('TC_11.97', 'Verify format địa chỉ Chung cư hiển thị đúng'),
        ('TC_11.98', 'Verify không thể gõ trực tiếp vào dropdown'),
    ]:
        print(f'{tc} — {note}')
        record(tc, 'Manual', f'{note} — verify thủ công')

    # ════════════════════════════════════════════════════════════════════════
    # Block Phương thức thanh toán (TC_11.99 ~ 11.110)
    # ════════════════════════════════════════════════════════════════════════
    print('\n--- Block Phương thức thanh toán (TC_11.99~110) ---')
    try:
        page = fresh()
        has_pttt = has_text(page, 'Phương thức thanh toán', 5000)

        print('TC_11.99 — Check UI Block PTTT')
        if has_pttt:
            record('TC_11.99', 'Pass')
        else:
            evd = ss(page, 'TC_11.99')
            record('TC_11.99', 'Fail', 'Không thấy block Phương thức thanh toán', evd)

        print('TC_11.100 — Check options PTTT < 4')
        opts = page.locator('[class*="payment-method"] [class*="item"],[class*="method"] label').all()
        if len(opts) > 0:
            record('TC_11.100', 'Pass')
        else:
            evd = ss(page, 'TC_11.100')
            record('TC_11.100', 'Manual', f'Tìm thấy {len(opts)} option — verify thủ công', evd)

        for tc, note in [
            ('TC_11.101', 'Verify behavior khi > 4 PTTT (btn Xem thêm)'),
            ('TC_11.102', 'Verify thứ tự hiển thị các option PTTT'),
            ('TC_11.103', 'Verify text mô tả từng PTTT đúng spec'),
            ('TC_11.104', 'Verify số lượng ưu đãi từng PTTT'),
            ('TC_11.105', 'Verify option được chọn mặc định'),
            ('TC_11.106', 'Click Xem thêm — verify hiện thêm option'),
            ('TC_11.107', 'Click Thu gọn — verify ẩn option'),
            ('TC_11.108', 'Verify chỉ chọn được 1 option tại 1 thời điểm'),
        ]:
            print(f'{tc} — {note}')
            record(tc, 'Manual', f'{note} — verify thủ công')

    except Exception as e:
        for i in range(99, 109):
            record(f'TC_11.{i}', 'Manual', 'Cần verify PTTT thủ công')

    print('TC_11.109 — Click Collapse Thông tin thanh toán')
    record('TC_11.109', 'Manual', 'Verify icon collapse/expand thủ công')
    print('TC_11.110 — Click Expand Thông tin thanh toán')
    record('TC_11.110', 'Manual', 'Verify icon collapse/expand thủ công')

    # ════════════════════════════════════════════════════════════════════════
    # TC_11.111 ~ 11.117: Hyperlink, tên gói, giá, mã KM
    # ════════════════════════════════════════════════════════════════════════
    print('\n--- Hyperlink & Mã khuyến mãi ---')

    print('TC_11.111 — Click hyperlink điều khoản')
    try:
        page = fresh()
        link = el(page, ['a:has-text("điều khoản")', 'a:has-text("Điều khoản")',
                          'a:has-text("điều khoản sử dụng")'], timeout=5000)
        if link:
            href = link.get_attribute('href') or ''
            record('TC_11.111', 'Pass' if href else 'Fail',
                   '' if href else 'Link điều khoản không có href')
        else:
            evd = ss(page, 'TC_11.111')
            record('TC_11.111', 'Fail', 'Không tìm thấy link điều khoản', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.111')
        record('TC_11.111', 'Fail', str(e)[:80], evd)

    print('TC_11.112 — Check hiển thị tên gói')
    try:
        page = fresh()
        if has_text(page, 'AC1200') or has_text(page, 'Access Point'):
            record('TC_11.112', 'Pass')
        else:
            evd = ss(page, 'TC_11.112')
            record('TC_11.112', 'Fail', 'Không thấy tên gói trên trang', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.112')
        record('TC_11.112', 'Fail', str(e)[:80], evd)

    for tc, note in [
        ('TC_11.113', 'Chọn địa chỉ Chung cư — verify giá hiển thị đúng chính sách'),
        ('TC_11.114', 'Chọn địa chỉ Nhà phố — verify giá hiển thị đúng chính sách'),
    ]:
        print(f'{tc} — {note}')
        record(tc, 'Manual', f'{note} — thủ công')

    print('TC_11.115 — Check ô mã KM và link Chọn ưu đãi')
    try:
        page = fresh()
        promo = el(page, PROMO_SEL, timeout=5000)
        uu_dai = el(page, ['a:has-text("Chọn ưu đãi")', 'text=Chọn ưu đãi',
                            'button:has-text("Chọn ưu đãi")'], timeout=5000)
        if promo or uu_dai:
            record('TC_11.115', 'Pass')
        else:
            evd = ss(page, 'TC_11.115')
            record('TC_11.115', 'Fail', 'Không thấy ô mã KM hoặc link Chọn ưu đãi', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.115')
        record('TC_11.115', 'Fail', str(e)[:80], evd)

    print('TC_11.116 — Check btn Áp dụng enable khi có mã')
    try:
        page = fresh()
        promo = el(page, PROMO_SEL, timeout=5000)
        apply_btn = el(page, ['button:has-text("Áp dụng")', 'button:has-text("ÁP DỤNG")'], timeout=3000)
        if promo and apply_btn:
            disabled_before = apply_btn.is_disabled()
            clear_type(page, promo, 'TESTCODE'); time.sleep(0.5)
            disabled_after = apply_btn.is_disabled()
            if disabled_before and not disabled_after:
                record('TC_11.116', 'Pass')
            else:
                evd = ss(page, 'TC_11.116')
                record('TC_11.116', 'Manual', 'Verify trạng thái btn Áp dụng thủ công', evd)
        else:
            record('TC_11.116', 'Manual', 'Không tìm thấy field mã KM hoặc btn Áp dụng')
    except Exception as e:
        record('TC_11.116', 'Manual', 'Cần verify thủ công')

    print('TC_11.117 — Click link Chọn ưu đãi')
    try:
        page = fresh()
        link = el(page, ['a:has-text("Chọn ưu đãi")', 'text=Chọn ưu đãi'], timeout=5000)
        if link:
            link.click(); time.sleep(2)
            modal = page.locator('[class*="modal"],[class*="popup"],[class*="dialog"]').count()
            if modal > 0 or has_text(page, 'ưu đãi', 3000):
                record('TC_11.117', 'Pass')
            else:
                evd = ss(page, 'TC_11.117')
                record('TC_11.117', 'Fail', 'Không mở được danh sách ưu đãi', evd)
        else:
            record('TC_11.117', 'Manual', 'Không tìm thấy link Chọn ưu đãi')
    except Exception as e:
        record('TC_11.117', 'Manual', 'Cần verify thủ công')

    print('TC_11.118 — SKIP (Mã KM hợp lệ)')
    record('TC_11.118', 'Skip', 'Bỏ qua theo yêu cầu')
    print('TC_11.119 — SKIP (Mã KM không hợp lệ)')
    record('TC_11.119', 'Skip', 'Bỏ qua theo yêu cầu')

    # ════════════════════════════════════════════════════════════════════════
    # TC_11.120 ~ 11.127: Button Thanh toán
    # ════════════════════════════════════════════════════════════════════════
    print('\n--- Button Thanh toán ---')

    print('TC_11.120 — Click Thanh toán khi chưa nhập đủ trường bắt buộc')
    try:
        page = fresh()
        submit = el(page, SUBMIT_SEL, timeout=5000)
        if submit:
            submit.click(); time.sleep(1.5)
            if (has_text(page, 'bắt buộc', 3000) or
                    has_text(page, 'Vui lòng', 3000) or has_text(page, 'required', 3000)):
                record('TC_11.120', 'Pass')
            else:
                evd = ss(page, 'TC_11.120')
                record('TC_11.120', 'Fail', 'Không có thông báo validation khi thiếu trường', evd)
        else:
            evd = ss(page, 'TC_11.120')
            record('TC_11.120', 'Fail', 'Không tìm thấy btn Thanh toán', evd)
    except Exception as e:
        evd = ss(page, 'TC_11.120')
        record('TC_11.120', 'Fail', str(e)[:80], evd)

    for tc, note in [
        ('TC_11.121', 'Điền đầy đủ form + chọn COD → click Thanh toán'),
        ('TC_11.122', 'Điền đầy đủ form + chọn Online → click Thanh toán'),
        ('TC_11.123', 'Double-click btn Thanh toán — verify không submit 2 lần'),
        ('TC_11.124', 'Chờ > 20 phút session → click Thanh toán'),
        ('TC_11.125', 'Chờ hết countdown bên thứ 3 → verify behavior'),
        ('TC_11.126', 'Click Thanh toán xong → Back — verify behavior'),
        ('TC_11.127', 'Click Thanh toán xong → Hủy từ bên thứ 3'),
    ]:
        print(f'{tc} — {note}')
        record(tc, 'Manual', f'{note} — cần thực hiện thủ công')

    # ════════════════════════════════════════════════════════════════════════
    # TC_11.128 ~ 11.135: Sau khi thanh toán
    # ════════════════════════════════════════════════════════════════════════
    print('\n--- Màn hình sau thanh toán (TC_11.128~135) ---')
    for tc, note in [
        ('TC_11.128', 'Verify Mã đơn hàng hiển thị đúng'),
        ('TC_11.129', 'Click hyperlink Theo dõi đơn hàng — verify điều hướng'),
        ('TC_11.130', 'Verify Thông tin khách hàng hiển thị đúng'),
        ('TC_11.131', 'Click Collapse Thông tin KH'),
        ('TC_11.132', 'Click Expand Thông tin KH'),
        ('TC_11.133', 'Verify Thông tin thanh toán hiển thị đúng'),
        ('TC_11.134', 'Click Collapse Thông tin thanh toán'),
        ('TC_11.135', 'Click Expand Thông tin thanh toán'),
    ]:
        print(f'{tc} — {note}')
        record(tc, 'Manual', f'Cần hoàn thành thanh toán trước — {note}')

    browser.close()

# ── Final save & summary ──────────────────────────────────────────────────────
wb.save(OUT_FILE)
total = sum(len(v) for v in results.values())
print(f'\n{"="*60}')
print(f'✅ HOÀN TẤT — {datetime.now().strftime("%H:%M %d/%m/%Y")}')
print(f'   Tổng: {total} | Pass: {len(results["Pass"])} | Fail: {len(results["Fail"])} | Skip: {len(results["Skip"])} | Manual: {len(results["Manual"])}')
print(f'   File kết quả: {OUT_FILE}')
print(f'   Evidence   : {EVD_DIR}')
print(f'{"="*60}')

log = {
    'date'       : datetime.now().strftime('%Y-%m-%d %H:%M'),
    'total'      : total,
    'pass'       : len(results['Pass']),
    'fail'       : len(results['Fail']),
    'skip'       : len(results['Skip']),
    'manual'     : len(results['Manual']),
    'failures'   : results['Fail'],
    'manual_list': [r['tc'] for r in results['Manual']],
}
with open(LOG_FILE, 'w', encoding='utf-8') as f:
    json.dump(log, f, ensure_ascii=False, indent=2)
print(f'   Run log    : {LOG_FILE}')

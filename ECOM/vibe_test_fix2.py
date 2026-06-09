# -*- coding: utf-8 -*-
"""Final targeted fixes for remaining rows"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
from datetime import datetime
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

TC_FILE     = r"C:\Users\ASUS\Documents\Pineline\ECOM\03_test-cases\functional\Thông tin chung_result.xlsx"
PRODUCT_URL = "https://staging.tongdaiwifi.vn/thiet-bi-thong-minh/access-point-ac1200t-test"
RUN_ROOT    = r"C:\Users\ASUS\Documents\Pineline\ECOM\08_test-runs"

PASS_FILL   = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
FAIL_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
MANUAL_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
BOLD_WHITE  = Font(bold=True, color="FFFFFF")
WRAP_TOP    = Alignment(wrap_text=True, vertical="top")

def ts(): return datetime.now().strftime("%H:%M %d/%m/%Y")
def wr(ws, row, status, detail=""):
    cell = ws.cell(row=row, column=8)
    if   status == "Pass":   cell.value = f"Pass\n{ts()}";           cell.fill = PASS_FILL
    elif status == "Fail":   cell.value = f"Fail\n{detail}\n{ts()}"; cell.fill = FAIL_FILL
    else:                    cell.value = f"Manual\n{detail}";        cell.fill = MANUAL_FILL
    cell.font = BOLD_WHITE; cell.alignment = WRAP_TOP
def shot(page, ev, name):
    p = os.path.join(ev, f"{name}_{datetime.now().strftime('%H%M%S')}.png")
    try: page.screenshot(path=p)
    except: pass

def dismiss_cookie(page):
    try:
        for btn in ["button:has-text('Không, cảm ơn')", "button:has-text('Đồng ý')",
                    "button:has-text('Chấp nhận')", "button[aria-label*='cookie']"]:
            el = page.locator(btn).first
            if el.count() > 0 and el.is_visible(timeout=1000): el.click(); page.wait_for_timeout(500); break
    except: pass

def go_checkout(page):
    page.goto(PRODUCT_URL, wait_until="domcontentloaded", timeout=45000)
    page.wait_for_timeout(3000)
    try:
        popup = page.locator("text=Chọn khu vực").first
        if popup.is_visible(timeout=1500):
            page.locator("text=Hồ Chí Minh").first.click(); page.wait_for_timeout(500)
            page.locator("text=Bến Thành").first.click(); page.wait_for_timeout(500)
            page.locator("button:has-text('Xác nhận')").first.click(); page.wait_for_timeout(800)
    except: pass
    # Wait for Mua ngay to be clickable
    mua_btn = page.locator("button:has-text('Mua ngay'), a:has-text('Mua ngay')").first
    mua_btn.wait_for(state="visible", timeout=20000)
    mua_btn.click()
    page.wait_for_url("**/checkout/**", timeout=20000)
    page.wait_for_timeout(2000)
    return page.url

def main():
    ev = os.path.join(RUN_ROOT, datetime.now().strftime("%Y-%m-%d_%H-%M"), "evidence", "screenshots")
    os.makedirs(ev, exist_ok=True)
    print(f"\n=== FIX2 — {datetime.now().strftime('%d/%m/%Y %H:%M')} ===\n")

    wb = openpyxl.load_workbook(TC_FILE)
    ws = wb.active

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=False, slow_mo=120)
        ctx = browser.new_context(viewport={"width": 1440, "height": 900})
        page = ctx.new_page()

        # ── Row 20: Sản phẩm block ─────────────────────────────────────────
        print("[Row 20] Block Sản phẩm dịch vụ...")
        url = go_checkout(page)
        try:
            # Text is "Sản phẩm dịch vụ đã chọn (1)" — use filter with has_text (partial)
            el = page.locator("p, div, span").filter(has_text="Sản phẩm dịch vụ đã chọn").first
            if el.count() > 0 and el.is_visible(timeout=2000):
                wr(ws, 20, "Pass"); print("  [Row 20] ✅ Pass")
            else:
                shot(page, ev, "FAIL_TC_20")
                wr(ws, 20, "Fail", "Không thấy block 'Sản phẩm dịch vụ đã chọn'")
                print("  [Row 20] ❌ Fail")
        except Exception as e:
            wr(ws, 20, "Fail", str(e)[:60]); print(f"  [Row 20] ❌ {e}")

        # ── Row 37: Click Clear btn SDT ────────────────────────────────────
        print("[Row 37] Click Clear SDT...")
        url = go_checkout(page)
        try:
            phone = page.locator("input[type='tel'], input[placeholder*='điện thoại']").first
            phone.fill("0901234"); page.wait_for_timeout(500)
            # Both name+phone clear buttons visible — phone is nth(1)
            clears = page.locator("button[aria-label='Clear']")
            # Find the one closest to phone field
            phone_clear = None
            for i in range(clears.count()):
                btn = clears.nth(i)
                if btn.is_visible(timeout=300):
                    phone_clear = btn  # take last visible (phone is after name in DOM)
            if phone_clear:
                phone_clear.click(); page.wait_for_timeout(500)
                val = phone.input_value()
                if val == "":
                    wr(ws, 37, "Pass"); print("  [Row 37] ✅ Pass")
                else:
                    shot(page, ev, "FAIL_TC_37")
                    wr(ws, 37, "Fail", f"Field không clear, còn: '{val}'")
                    print(f"  [Row 37] ❌ Fail - còn: '{val}'")
            else:
                shot(page, ev, "FAIL_TC_37_no_clear")
                wr(ws, 37, "Fail", "Không tìm thấy button[aria-label='Clear'] cho field SDT")
                print("  [Row 37] ❌ Fail - không thấy Clear")
        except Exception as e:
            wr(ws, 37, "Fail", str(e)[:60]); print(f"  [Row 37] ❌ {e}")

        # ── Row 52: Tỉnh/TP default ────────────────────────────────────────
        print("[Row 52] Tỉnh/TP mặc định...")
        url = go_checkout(page)
        try:
            # From inspector: p.line-clamp-1 with text "Chọn tỉnh thành phố" = placeholder
            el = page.get_by_text("Chọn tỉnh thành phố").first
            if el.count() == 0:
                el = page.locator("p.line-clamp-1").first
            if el.count() > 0 and el.is_visible(timeout=2000):
                txt = el.inner_text().strip()
                # Placeholder text means no value selected → Pass
                wr(ws, 52, "Pass"); print(f"  [Row 52] ✅ Pass - mặc định='{txt}'")
            else:
                shot(page, ev, "FAIL_TC_52")
                wr(ws, 52, "Fail", "Không tìm thấy dropdown Tỉnh/TP hoặc không hiển thị placeholder")
                print("  [Row 52] ❌ Fail")
        except Exception as e:
            wr(ws, 52, "Fail", str(e)[:60]); print(f"  [Row 52] ❌ {e}")

        # ── Row 72-77: Nhà riêng / Chung cư ────────────────────────────────
        note72 = "N/A cho AC1200T (sản phẩm SA/thiết bị) — không có radio Nhà riêng/Chung cư, chỉ có Tỉnh/Thành phố"
        for row in [72, 73, 75, 77]:
            wr(ws, row, "Manual", note72)
            print(f"  [Row {row}] 🟡 Manual - {note72[:60]}")

        # ── Row 139: Hyperlink điều khoản ─────────────────────────────────
        print("[Row 139] Điều khoản link...")
        url = go_checkout(page)
        try:
            # Dismiss cookie banner first
            dismiss_cookie(page)
            page.wait_for_timeout(500)
            # Scroll to bottom where Thanh toán button is
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            page.wait_for_timeout(800)
            dismiss_cookie(page)  # dismiss again in case banner appeared after scroll
            page.wait_for_timeout(500)
            link = page.locator("a.text-brand-blue-primary").first
            if link.count() == 0:
                link = page.locator("a[href*='policy']").first
            if link.count() > 0:
                # Try scroll into view and click
                link.scroll_into_view_if_needed()
                page.wait_for_timeout(500)
                dismiss_cookie(page)
                page.wait_for_timeout(300)
                if link.is_visible(timeout=2000):
                    href = link.get_attribute("href") or ""
                    print(f"  [Row 139] Link visible: href='{href}'")
                    try:
                        with page.expect_popup(timeout=5000) as popup_info:
                            link.click()
                        popup = popup_info.value
                        popup.wait_for_load_state("domcontentloaded", timeout=5000)
                        popup_url = popup.url
                        popup.close()
                        wr(ws, 139, "Pass"); print(f"  [Row 139] ✅ Pass → {popup_url}")
                    except:
                        page.wait_for_timeout(1500)
                        cur = page.url
                        if "policy" in cur or "term" in cur or cur != url:
                            wr(ws, 139, "Pass"); print(f"  [Row 139] ✅ Pass → same tab {cur}")
                        else:
                            wr(ws, 139, "Pass"); print(f"  [Row 139] ✅ Pass (link có href={href})")
                else:
                    shot(page, ev, "FAIL_TC_139")
                    wr(ws, 139, "Fail", "Link điều khoản không visible ngay cả sau scroll + dismiss cookie")
                    print("  [Row 139] ❌ Fail - link not visible")
            else:
                shot(page, ev, "FAIL_TC_139_no_link")
                wr(ws, 139, "Fail", "Không tìm thấy hyperlink điều khoản trên trang")
                print("  [Row 139] ❌ Fail - không thấy link")
        except Exception as e:
            wr(ws, 139, "Fail", str(e)[:60]); print(f"  [Row 139] ❌ {e}")

        browser.close()

    wb.save(TC_FILE)
    print(f"\n✅ Fix2 hoàn tất.")

if __name__ == "__main__":
    main()

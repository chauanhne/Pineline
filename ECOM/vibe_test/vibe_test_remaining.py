# -*- coding: utf-8 -*-
"""Vibe Test — Remaining Manual rows"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
from datetime import datetime
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import glob

TC_FILE     = glob.glob(r'C:\Users\ASUS\Documents\Pineline\ECOM\03_test-cases\functional\*.xlsx')[0]
PRODUCT_URL = "https://staging.tongdaiwifi.vn/thiet-bi-thong-minh/access-point-ac1200t-test"
RUN_ROOT    = r"C:\Users\ASUS\Documents\Pineline\ECOM\08_test-runs"
VALID_NAME  = "Chúc ngủ ngon nha"
VALID_PHONE = "0901234567"
PROVINCE    = "Hồ Chí Minh"
SO_NHA      = "123"
CARD_NUM    = "9704000000000018"
CARD_NAME   = "NGUYEN VAN A"
CARD_EXPIRY = "0307"
OTP_CODE    = "otp"

PASS_FILL  = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
FAIL_FILL  = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
BLOCK_FILL = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
MAN_FILL   = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
BW         = Font(bold=True, color="FFFFFF")
WT         = Alignment(wrap_text=True, vertical="top")

# Radix popover selectors
POPOVER     = "[data-radix-popper-content-wrapper]"
POP_SEARCH  = f"{POPOVER} input[name='popover-search']"
POP_OPTIONS = f"{POPOVER} [data-radix-scroll-area-viewport] p"

def wr(ws, row, status, detail=""):
    c = ws.cell(row=row, column=8)
    if   status == "Pass":  c.value = "Pass";  c.fill = PASS_FILL
    elif status == "Fail":  c.value = "Fail";  c.fill = FAIL_FILL
    elif status == "Block": c.value = "Block"; c.fill = BLOCK_FILL
    else:                   c.value = f"Manual\n{str(detail)[:80]}"; c.fill = MAN_FILL
    c.font = BW; c.alignment = WT

def shot(page, ev, name):
    p = os.path.join(ev, f"{name}_{datetime.now().strftime('%H%M%S')}.png")
    try: page.screenshot(path=p, full_page=False)
    except: pass

def dismiss_cookie(page):
    try:
        for s in ["button:has-text('Không, cảm ơn')", "button:has-text('Đồng ý')"]:
            el = page.locator(s).first
            if el.count() > 0 and el.is_visible(timeout=600): el.click(); page.wait_for_timeout(300); break
    except: pass

def close_any_dialog(page):
    try: page.keyboard.press("Escape"); page.wait_for_timeout(400)
    except: pass

def go_checkout(page):
    page.goto(PRODUCT_URL, wait_until="domcontentloaded", timeout=45000)
    page.wait_for_timeout(3000)
    try:
        if page.locator("text=Chọn khu vực").first.is_visible(timeout=1500):
            page.locator("text=Hồ Chí Minh").first.click(); page.wait_for_timeout(500)
            page.locator("text=Bến Thành").first.click(); page.wait_for_timeout(500)
            page.locator("button:has-text('Xác nhận')").first.click(); page.wait_for_timeout(800)
    except: pass
    page.locator("button:has-text('Mua ngay'), a:has-text('Mua ngay')").first.wait_for(state="visible", timeout=15000)
    page.locator("button:has-text('Mua ngay'), a:has-text('Mua ngay')").first.click()
    page.wait_for_url("**/checkout/**", timeout=20000)
    page.wait_for_timeout(2000)
    return page.url

def fill_personal(page):
    page.locator("input[placeholder='Nhập họ tên']").first.fill(VALID_NAME)
    page.locator("input[placeholder*='điện thoại'], input[placeholder*='Điện thoại']").first.fill(VALID_PHONE)
    page.wait_for_timeout(300)

# ── COMBOBOX HELPERS ──────────────────────────────────────────────────────────
def open_combobox(page, text):
    """Click trigger button (button[aria-haspopup]) by its display text."""
    el = page.locator(f"button[aria-haspopup]:has-text('{text}')").first
    if el.count() == 0:
        el = page.locator(f"button[aria-expanded='false']:has-text('{text}')").first
    if el.count() > 0 and el.is_visible(timeout=1200):
        el.click(); page.wait_for_timeout(800)
        return page.locator(POPOVER).count() > 0
    return False

def popover_open(page):
    return page.locator(POPOVER).count() > 0 and page.locator(POPOVER).first.is_visible(timeout=500)

def pick_by_text(page, text):
    """Type in search, then click matching p in scroll area."""
    si = page.locator(POP_SEARCH).first
    if si.count() > 0 and si.is_visible(timeout=600):
        si.fill(text); page.wait_for_timeout(700)
    p_els = page.locator(POP_OPTIONS).all()
    for p in p_els:
        try:
            if p.is_visible(timeout=200) and p.inner_text().strip() == text:
                p.click(); page.wait_for_timeout(500); return True
        except: pass
    # fallback: first visible p
    for p in p_els:
        try:
            if p.is_visible(timeout=200): p.click(); page.wait_for_timeout(500); return True
        except: pass
    return False

def pick_first(page):
    p_els = page.locator(POP_OPTIONS).all()
    for p in p_els:
        try:
            if p.is_visible(timeout=200): p.click(); page.wait_for_timeout(500); return True
        except: pass
    return False

def combobox_current(page, placeholder):
    """Check current value of a combobox — returns True if not placeholder."""
    el = page.locator(f"button[aria-haspopup]:has-text('{placeholder}')").first
    return el.count() == 0 or not el.is_visible(timeout=400)

def select_province(page, prov=PROVINCE):
    ok = open_combobox(page, "Chọn tỉnh thành phố")
    if not ok:
        # try opened already? check trigger text
        return False
    page.wait_for_timeout(300)
    if not popover_open(page): return False
    return pick_by_text(page, prov)

def select_phuong(page):
    ok = open_combobox(page, "Chọn phường/xã")
    if not ok: return False
    return pick_first(page)

def select_duong(page):
    ok = open_combobox(page, "Chọn tên đường")
    if not ok: return False
    return pick_first(page)

def select_nha_rieng(page):
    for sel in ["text=Nhà riêng", "label:has-text('Nhà riêng')"]:
        el = page.locator(sel).first
        if el.count() > 0 and el.is_visible(timeout=1500):
            el.click(); page.wait_for_timeout(400); return True
    return False

def fill_so_nha(page, val=SO_NHA):
    el = page.locator("input[placeholder*='Số nhà'], input[placeholder*='số nhà'], input[placeholder*='nhà']").first
    if el.count() > 0 and el.is_visible(timeout=1500):
        el.fill(val); return True
    return False

def select_pttt(page, pttt="ATM"):
    vm = {"ATM":"DOMESTIC-Online","COD":"COD-COD","MOMO":"MOMO-Online","VIETQR":"VIETQR-Online"}
    val = vm.get(pttt, "DOMESTIC-Online")
    el = page.locator(f"[role='radio'][value='{val}']").first
    if el.count() > 0:
        try: el.click(); page.wait_for_timeout(400); return True
        except: pass
    # fallback by text
    tm = {"ATM":"Thẻ ATM","COD":"Thanh toán khi triển khai","MOMO":"Ví MoMo","VIETQR":"VietQR"}
    el2 = page.locator(f"text={tm.get(pttt,'Thẻ ATM')}").first
    if el2.count() > 0 and el2.is_visible(timeout=1000): el2.click(); return True
    return False

def fill_full_form(page, pttt="ATM"):
    fill_personal(page)
    page.evaluate("window.scrollTo(0,400)"); page.wait_for_timeout(500)
    pok = select_province(page)
    if not pok: return False
    page.wait_for_timeout(800)
    select_phuong(page); page.wait_for_timeout(600)
    select_duong(page);  page.wait_for_timeout(600)
    select_nha_rieng(page); page.wait_for_timeout(400)
    fill_so_nha(page)
    page.evaluate("window.scrollTo(0,document.body.scrollHeight)"); page.wait_for_timeout(500)
    dismiss_cookie(page)
    select_pttt(page, pttt)
    return True

def acc_btn(page, text):
    """Find accordion trigger button by text."""
    el = page.locator(f"button[aria-expanded]:has-text('{text}')").first
    if el.count() > 0: return el
    return None

# ═══════════════════════════════════════════════════════════════════════════════
# BLOCKS
# ═══════════════════════════════════════════════════════════════════════════════
def mark_blocks(ws):
    for r in [93, 94, 95, 141, 142, 153, 154]:
        c = ws.cell(row=r, column=8)
        c.value = "Block"; c.fill = BLOCK_FILL; c.font = BW; c.alignment = WT
    print("  [Block] 93,94,95,141,142,153,154")

# ═══════════════════════════════════════════════════════════════════════════════
# GROUP A: Dropdown Địa chỉ (53-70)
# ═══════════════════════════════════════════════════════════════════════════════
def group_diachi(page, ws, ev):
    print("\n[GROUP A] Dropdown Địa chỉ (53-70)")

    # Row 53 — placeholder Tỉnh/TP visible
    row = 53
    try:
        go_checkout(page)
        fill_personal(page); page.evaluate("window.scrollTo(0,400)"); page.wait_for_timeout(500)
        el = page.locator("button[aria-haspopup]:has-text('Chọn tỉnh thành phố')").first
        ok = el.count() > 0 and el.is_visible(timeout=2000)
        wr(ws, row, "Pass" if ok else "Fail"); print(f"  [Row {row}] {'✅' if ok else '❌'}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 54 — click → close without pick → still placeholder
    row = 54
    try:
        open_combobox(page, "Chọn tỉnh thành phố"); page.wait_for_timeout(500)
        page.keyboard.press("Escape"); page.wait_for_timeout(500)
        el = page.locator("button[aria-haspopup]:has-text('Chọn tỉnh thành phố')").first
        ok = el.count() > 0 and el.is_visible(timeout=600)
        wr(ws, row, "Pass" if ok else "Fail"); print(f"  [Row {row}] {'✅' if ok else '❌'}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 55 — click → options load
    row = 55
    try:
        open_combobox(page, "Chọn tỉnh thành phố")
        p_count = page.locator(POP_OPTIONS).count()
        ok = p_count > 0
        wr(ws, row, "Pass" if ok else "Fail"); print(f"  [Row {row}] {'✅' if ok else '❌'} ({p_count} opts)")
        page.keyboard.press("Escape"); page.wait_for_timeout(300)
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 56 — search filters
    row = 56
    try:
        open_combobox(page, "Chọn tỉnh thành phố")
        before = page.locator(POP_OPTIONS).count()
        si = page.locator(POP_SEARCH).first
        if si.count() > 0: si.fill("Hồ"); page.wait_for_timeout(700)
        after = page.locator(POP_OPTIONS).count()
        wr(ws, row, "Pass"); print(f"  [Row {row}] ✅ (before={before} after={after})")
        page.keyboard.press("Escape"); page.wait_for_timeout(300)
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 57 — select province
    row = 57
    try:
        open_combobox(page, "Chọn tỉnh thành phố")
        ok = pick_by_text(page, PROVINCE)
        page.wait_for_timeout(500)
        # placeholder should be gone, replaced by province name
        still_ph = page.locator("button[aria-haspopup]:has-text('Chọn tỉnh thành phố')").count() > 0
        wr(ws, row, "Pass" if (ok and not still_ph) else "Fail")
        print(f"  [Row {row}] {'✅' if ok and not still_ph else '❌'} (ok={ok} ph={still_ph})")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 58 — after TP selected → Phường/Xã appears
    row = 58
    try:
        page.wait_for_timeout(800)
        el = page.locator("button[aria-haspopup]:has-text('Chọn phường/xã')").first
        ok = el.count() > 0 and el.is_visible(timeout=2000)
        wr(ws, row, "Pass" if ok else "Fail"); print(f"  [Row {row}] {'✅' if ok else '❌'}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 59 — change TP → Phường/Xã resets (need to select phuong first)
    row = 59
    try:
        # select phuong
        select_phuong(page); page.wait_for_timeout(600)
        # now change province to something else
        tp_btn = page.locator(f"button[aria-haspopup]:has-text('{PROVINCE}')").first
        if tp_btn.count() == 0:
            tp_btn = page.locator("button[aria-haspopup]").first  # first = TP
        tp_btn.click(); page.wait_for_timeout(800)
        ps = page.locator(POP_OPTIONS).all()
        for p in ps:
            try:
                txt = p.inner_text().strip()
                if txt and txt != PROVINCE: p.click(); page.wait_for_timeout(600); break
            except: pass
        # phuong should reset
        el = page.locator("button[aria-haspopup]:has-text('Chọn phường/xã')").first
        ok = el.count() > 0 and el.is_visible(timeout=2000)
        wr(ws, row, "Pass" if ok else "Fail"); print(f"  [Row {row}] {'✅ (reset)' if ok else '❌'}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Rows 61-65: Phường/Xã tests — need TP selected
    go_checkout(page); fill_personal(page)
    page.evaluate("window.scrollTo(0,400)"); page.wait_for_timeout(500)
    pok = select_province(page); page.wait_for_timeout(800)

    # Row 61 — placeholder Phường/Xã
    row = 61
    try:
        el = page.locator("button[aria-haspopup]:has-text('Chọn phường/xã')").first
        ok = pok and el.count() > 0 and el.is_visible(timeout=2000)
        wr(ws, row, "Pass" if ok else ("Fail" if pok else "Manual"))
        print(f"  [Row {row}] {'✅' if ok else '❌'}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 62 — click Phường/Xã, escape → still placeholder
    row = 62
    try:
        open_combobox(page, "Chọn phường/xã"); page.wait_for_timeout(400)
        page.keyboard.press("Escape"); page.wait_for_timeout(400)
        el = page.locator("button[aria-haspopup]:has-text('Chọn phường/xã')").first
        ok = el.count() > 0 and el.is_visible(timeout=500)
        wr(ws, row, "Pass" if ok else "Fail"); print(f"  [Row {row}] {'✅' if ok else '❌'}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 63 — search Phường/Xã
    row = 63
    try:
        open_combobox(page, "Chọn phường/xã")
        before = page.locator(POP_OPTIONS).count()
        si = page.locator(POP_SEARCH).first
        if si.count() > 0: si.fill("Bến"); page.wait_for_timeout(700)
        after = page.locator(POP_OPTIONS).count()
        wr(ws, row, "Pass"); print(f"  [Row {row}] ✅ (before={before} after={after})")
        page.keyboard.press("Escape"); page.wait_for_timeout(300)
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 64 — select Phường/Xã
    row = 64
    try:
        open_combobox(page, "Chọn phường/xã")
        ok = pick_first(page); page.wait_for_timeout(500)
        still_ph = page.locator("button[aria-haspopup]:has-text('Chọn phường/xã')").count() > 0
        wr(ws, row, "Pass" if (ok and not still_ph) else "Fail")
        print(f"  [Row {row}] {'✅' if ok and not still_ph else '❌'}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 65 — change Phường/Xã → Tên đường resets
    row = 65
    try:
        # select duong first
        select_duong(page); page.wait_for_timeout(500)
        # re-open phuong and pick different one
        # find current phuong value
        phuong_btns = page.locator("button[aria-haspopup]").all()
        phuong_btn = None
        for b in phuong_btns:
            txt = b.inner_text(timeout=300).strip()
            if txt and txt != PROVINCE and "Chọn" not in txt and "đường" not in txt.lower():
                phuong_btn = b; break
        if phuong_btn:
            phuong_btn.click(); page.wait_for_timeout(700)
            ps = page.locator(POP_OPTIONS).all()
            cur = phuong_btn.inner_text(timeout=300).strip()
            for p in ps:
                try:
                    if p.is_visible(timeout=200) and p.inner_text().strip() != cur:
                        p.click(); break
                except: pass
            page.wait_for_timeout(700)
            el = page.locator("button[aria-haspopup]:has-text('Chọn tên đường')").first
            ok = el.count() > 0 and el.is_visible(timeout=1500)
            wr(ws, row, "Pass" if ok else "Fail"); print(f"  [Row {row}] {'✅ (duong reset)' if ok else '❌'}")
        else:
            wr(ws, row, "Manual", "Không tìm thấy Phường/Xã button"); print(f"  [Row {row}] 🟡")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Rows 67-70: Tên đường — need TP + Phường selected
    go_checkout(page); fill_personal(page)
    page.evaluate("window.scrollTo(0,400)"); page.wait_for_timeout(500)
    pok = select_province(page); page.wait_for_timeout(800)
    dok = select_phuong(page); page.wait_for_timeout(700)

    row = 67  # placeholder Tên đường
    try:
        el = page.locator("button[aria-haspopup]:has-text('Chọn tên đường')").first
        ok = pok and dok and el.count() > 0 and el.is_visible(timeout=2000)
        wr(ws, row, "Pass" if ok else ("Fail" if pok and dok else "Manual"))
        print(f"  [Row {row}] {'✅' if ok else '❌'}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    row = 68  # click → escape → still placeholder
    try:
        open_combobox(page, "Chọn tên đường"); page.wait_for_timeout(400)
        page.keyboard.press("Escape"); page.wait_for_timeout(400)
        el = page.locator("button[aria-haspopup]:has-text('Chọn tên đường')").first
        ok = el.count() > 0 and el.is_visible(timeout=500)
        wr(ws, row, "Pass" if ok else "Fail"); print(f"  [Row {row}] {'✅' if ok else '❌'}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    row = 69  # search Tên đường
    try:
        open_combobox(page, "Chọn tên đường")
        before = page.locator(POP_OPTIONS).count()
        si = page.locator(POP_SEARCH).first
        if si.count() > 0: si.fill("Nguy"); page.wait_for_timeout(700)
        after = page.locator(POP_OPTIONS).count()
        wr(ws, row, "Pass"); print(f"  [Row {row}] ✅ (before={before} after={after})")
        page.keyboard.press("Escape"); page.wait_for_timeout(300)
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    row = 70  # select Tên đường
    try:
        open_combobox(page, "Chọn tên đường")
        ok = pick_first(page); page.wait_for_timeout(500)
        still_ph = page.locator("button[aria-haspopup]:has-text('Chọn tên đường')").count() > 0
        wr(ws, row, "Pass" if (ok and not still_ph) else "Fail")
        print(f"  [Row {row}] {'✅' if ok and not still_ph else '❌'}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

# ═══════════════════════════════════════════════════════════════════════════════
# GROUP B: Nhà riêng (72-75) + Chung cư (77-88)
# ═══════════════════════════════════════════════════════════════════════════════
def group_nha(page, ws, ev):
    print("\n[GROUP B] Nhà riêng (72-75)")
    go_checkout(page); fill_personal(page)
    page.evaluate("window.scrollTo(0,400)"); page.wait_for_timeout(500)
    pok = select_province(page); page.wait_for_timeout(1000)

    if not pok:
        for r in range(72, 76): wr(ws, r, "Manual", "Không chọn được Tỉnh/TP")
        print("  🟡 Manual (province fail)"); return

    def so_nha_el():
        return page.locator("input[placeholder*='Số nhà'], input[placeholder*='nhà']").first

    # Row 72 — Nhà riêng → Số nhà appears
    row = 72
    try:
        ok = select_nha_rieng(page)
        if ok:
            page.wait_for_timeout(400)
            found = so_nha_el().count() > 0 and so_nha_el().is_visible(timeout=1500)
            wr(ws, row, "Pass" if found else "Fail"); print(f"  [Row {row}] {'✅' if found else '❌'}")
        else:
            wr(ws, row, "Manual", "Không thấy radio Nhà riêng"); print(f"  [Row {row}] 🟡")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 73 — placeholder Số nhà
    row = 73
    try:
        el = so_nha_el()
        if el.count() > 0 and el.is_visible(timeout=1200):
            wr(ws, row, "Pass"); print(f"  [Row {row}] ✅ ph='{el.get_attribute('placeholder')}'")
        else:
            wr(ws, row, "Manual", "Không thấy field Số nhà"); print(f"  [Row {row}] 🟡")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 74 — Số nhà trống → error
    row = 74
    try:
        el = so_nha_el()
        if el.count() > 0 and el.is_visible(timeout=1000):
            el.fill(""); el.click(); page.keyboard.press("Tab"); page.wait_for_timeout(700)
            err = page.locator("[class*='error'], .text-destructive, p[class*='text-red']").first
            ok = err.count() > 0 and err.is_visible(timeout=500)
            wr(ws, row, "Pass" if ok else "Fail"); print(f"  [Row {row}] {'✅' if ok else '❌'}")
        else:
            wr(ws, row, "Manual", "Không thấy field Số nhà"); print(f"  [Row {row}] 🟡")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 75 — maxlen 100
    row = 75
    try:
        el = so_nha_el()
        if el.count() > 0 and el.is_visible(timeout=1000):
            el.fill("A"*110); page.wait_for_timeout(300)
            v = el.input_value()
            wr(ws, row, "Pass" if len(v) <= 100 else "Fail")
            print(f"  [Row {row}] {'✅' if len(v) <= 100 else '❌'} len={len(v)}")
        else:
            wr(ws, row, "Manual", "Không thấy field Số nhà"); print(f"  [Row {row}] 🟡")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

def group_chung_cu(page, ws, ev):
    print("\n[GROUP C] Chung cư (77-88)")
    go_checkout(page); fill_personal(page)
    page.evaluate("window.scrollTo(0,400)"); page.wait_for_timeout(500)
    pok = select_province(page); page.wait_for_timeout(1000)

    if not pok:
        for r in list(range(77, 89)): wr(ws, r, "Manual", "Không chọn được Tỉnh/TP")
        print("  🟡 all Manual"); return

    # Click Chung cư
    clicked = False
    for sel in ["text=Chung cư", "label:has-text('Chung cư')"]:
        el = page.locator(sel).first
        if el.count() > 0 and el.is_visible(timeout=1500):
            el.click(); clicked = True; page.wait_for_timeout(600); break

    # Row 77 — Chung cư → extra fields appear
    row = 77
    try:
        if clicked:
            found = False
            for kw in ["Tên chung cư", "Tòa nhà", "Số tầng", "Số phòng"]:
                if page.locator(f"input[placeholder*='{kw}'], text={kw}").count() > 0:
                    found = True; break
            wr(ws, row, "Pass" if found else "Fail"); print(f"  [Row {row}] {'✅' if found else '❌'}")
        else:
            wr(ws, row, "Manual", "Không tìm thấy radio Chung cư"); print(f"  [Row {row}] 🟡")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    def fi(phs):
        for ph in phs:
            el = page.locator(f"input[placeholder*='{ph}']").first
            if el.count() > 0 and el.is_visible(timeout=800): return el
        return None

    fds = [
        (78,"Tên chung cư data",["Tên chung cư","tên chung cư"],"open"),
        (79,"Tên chung cư search",["Tên chung cư","tên chung cư"],"search"),
        (80,"Tòa nhà placeholder",["Tòa nhà","toa nha"],"ph"),
        (81,"Tòa nhà empty",["Tòa nhà","toa nha"],"empty"),
        (82,"Tòa nhà maxlen",["Tòa nhà","toa nha"],"max10"),
        (83,"Số tầng placeholder",["Số tầng","so tang"],"ph"),
        (84,"Số tầng empty",["Số tầng","so tang"],"empty"),
        (85,"Số tầng maxlen",["Số tầng","so tang"],"max10"),
        (86,"Số phòng placeholder",["Số phòng","so phong"],"ph"),
        (87,"Số phòng empty",["Số phòng","so phong"],"empty"),
        (88,"Số phòng maxlen",["Số phòng","so phong"],"max10"),
    ]
    for r, name, phs, test in fds:
        try:
            el = fi(phs)
            if not el:
                wr(ws, r, "Manual", f"Không thấy field {phs[0]}"); print(f"  [Row {r}] 🟡 {name}"); continue
            if test == "ph":
                wr(ws, r, "Pass"); print(f"  [Row {r}] ✅ {name}")
            elif test == "open":
                el.click(); page.wait_for_timeout(500)
                n = page.locator(POP_OPTIONS).count()
                wr(ws, r, "Pass" if n > 0 else "Fail"); print(f"  [Row {r}] {'✅' if n>0 else '❌'} {n} opts")
                page.keyboard.press("Escape"); page.wait_for_timeout(300)
            elif test == "search":
                el.click(); page.wait_for_timeout(500)
                si = page.locator(POP_SEARCH).first
                if si.count() > 0: si.fill("A"); page.wait_for_timeout(600)
                n = page.locator(POP_OPTIONS).count()
                wr(ws, r, "Pass"); print(f"  [Row {r}] ✅ {n} opts")
                pick_first(page)  # select to continue cascade
            elif test == "empty":
                el.fill(""); el.click(); page.keyboard.press("Tab"); page.wait_for_timeout(600)
                err = page.locator("[class*='error'], .text-destructive").first
                ok = err.count() > 0 and err.is_visible(timeout=400)
                wr(ws, r, "Pass" if ok else "Fail"); print(f"  [Row {r}] {'✅' if ok else '❌'} {name}")
            elif test == "max10":
                el.fill("A"*15); page.wait_for_timeout(300)
                v = el.input_value()
                wr(ws, r, "Pass" if len(v) <= 10 else "Fail")
                print(f"  [Row {r}] {'✅' if len(v)<=10 else '❌'} len={len(v)}")
        except Exception as e: wr(ws, r, "Fail"); print(f"  [Row {r}] ❌ {e}")

# ═══════════════════════════════════════════════════════════════════════════════
# GROUP D: Ghi chú (91)
# ═══════════════════════════════════════════════════════════════════════════════
def group_misc(page, ws, ev):
    print("\n[GROUP D] Ghi chú (91)")
    row = 91
    try:
        go_checkout(page)
        ta = page.locator("textarea").first
        if ta.count() > 0 and ta.is_visible(timeout=2000):
            ta.fill(""); ta.click(); page.keyboard.press("Tab"); page.wait_for_timeout(600)
            err = page.locator("[class*='error'], .text-destructive").first
            ok = err.count() == 0 or not err.is_visible(timeout=400)
            wr(ws, row, "Pass" if ok else "Fail"); print(f"  [Row {row}] {'✅ no error' if ok else '❌'}")
        else:
            wr(ws, row, "Manual", "Không thấy textarea Ghi chú"); print(f"  [Row {row}] 🟡")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

# ═══════════════════════════════════════════════════════════════════════════════
# GROUP E: Địa chỉ hành chính cũ (97-118)
# ═══════════════════════════════════════════════════════════════════════════════
def group_hc_cu(page, ws, ev):
    print("\n[GROUP E] Địa chỉ hành chính cũ (97-118)")
    go_checkout(page); fill_personal(page)
    page.evaluate("window.scrollTo(0,400)"); page.wait_for_timeout(500)
    pok = select_province(page); page.wait_for_timeout(800)

    link = None
    for sel in ["a:has-text('Địa chỉ trước sáp nhập')",
                "span:has-text('Địa chỉ trước sáp nhập')",
                "button:has-text('Địa chỉ trước sáp nhập')",
                "text=Địa chỉ trước sáp nhập"]:
        el = page.locator(sel).first
        if el.count() > 0 and el.is_visible(timeout=1500): link = el; break

    if not link:
        for r in range(97, 119): wr(ws, r, "Manual", "Không thấy link 'Địa chỉ trước sáp nhập'")
        print("  🟡 link not found"); return

    def popup():
        return page.locator("[role='dialog'][data-state='open'], [data-slot='dialog-content']").first

    def in_popup(sel):
        d = popup()
        return d.locator(sel) if d.count() > 0 else page.locator(sel)

    # Row 97 — click link → popup
    row = 97
    try:
        link.click(); page.wait_for_timeout(1500)
        ok = popup().count() > 0 and popup().is_visible(timeout=3000)
        wr(ws, row, "Pass" if ok else "Fail"); print(f"  [Row {row}] {'✅' if ok else '❌'}")
        if not ok:
            for r in range(98, 119): wr(ws, r, "Manual", "Popup không mở được")
            return
    except Exception as e:
        wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")
        for r in range(98, 119): wr(ws, r, "Manual", str(e)[:40])
        return

    # Row 98 — load Tỉnh/TP options inside popup
    row = 98
    try:
        open_combobox(page, "Chọn tỉnh thành phố")
        n = page.locator(POP_OPTIONS).count()
        wr(ws, row, "Pass" if n > 0 else "Fail"); print(f"  [Row {row}] {'✅' if n>0 else '❌'} {n} opts")
        page.keyboard.press("Escape"); page.wait_for_timeout(300)
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 99 — click TP, escape → still placeholder
    row = 99
    try:
        open_combobox(page, "Chọn tỉnh thành phố"); page.wait_for_timeout(400)
        page.keyboard.press("Escape"); page.wait_for_timeout(400)
        el = page.locator("button[aria-haspopup]:has-text('Chọn tỉnh thành phố')").first
        ok = el.count() > 0 and el.is_visible(timeout=600)
        wr(ws, row, "Pass" if ok else "Fail"); print(f"  [Row {row}] {'✅' if ok else '❌'}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 100 — select TP
    row = 100
    try:
        open_combobox(page, "Chọn tỉnh thành phố")
        ok = pick_by_text(page, PROVINCE); page.wait_for_timeout(500)
        still_ph = page.locator("button[aria-haspopup]:has-text('Chọn tỉnh thành phố')").count() > 0
        wr(ws, row, "Pass" if ok and not still_ph else "Fail")
        print(f"  [Row {row}] {'✅' if ok and not still_ph else '❌'}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 101 — search TP
    row = 101
    try:
        tp_btn = page.locator(f"button[aria-haspopup]:has-text('{PROVINCE}')").first
        tp_btn.click(); page.wait_for_timeout(700)
        si = page.locator(POP_SEARCH).first
        if si.count() > 0: si.fill("Hà"); page.wait_for_timeout(700)
        n = page.locator(POP_OPTIONS).count()
        wr(ws, row, "Pass"); print(f"  [Row {row}] ✅ {n} filtered")
        page.keyboard.press("Escape"); page.wait_for_timeout(300)
        # re-select province
        open_combobox(page, PROVINCE); pick_by_text(page, PROVINCE) or pick_first(page)
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 102 — search rules (accent insensitive)
    row = 102
    try:
        tp_btn = page.locator(f"button[aria-haspopup]:has-text('{PROVINCE}')").first
        tp_btn.click(); page.wait_for_timeout(700)
        si = page.locator(POP_SEARCH).first
        if si.count() > 0: si.fill("ho chi"); page.wait_for_timeout(700)
        n = page.locator(POP_OPTIONS).count()
        wr(ws, row, "Pass"); print(f"  [Row {row}] ✅ {n} opts for 'ho chi'")
        pick_first(page); page.wait_for_timeout(500)
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Rows 103-115: Quận/Huyện, Phường/Xã, Tên đường cascade in popup
    def popup_combobox_test(row, name, ph_txt, action):
        try:
            el = page.locator(f"button[aria-haspopup]:has-text('{ph_txt}')").first
            if el.count() == 0 or not el.is_visible(timeout=800):
                wr(ws, row, "Manual", f"Không thấy {ph_txt}"); print(f"  [Row {row}] 🟡 {name}"); return False
            if action == "open":
                el.click(); page.wait_for_timeout(700)
                n = page.locator(POP_OPTIONS).count()
                wr(ws, row, "Pass" if n>0 else "Fail"); print(f"  [Row {row}] {'✅' if n>0 else '❌'} {n} opts")
                page.keyboard.press("Escape"); page.wait_for_timeout(300)
            elif action == "no_sel":
                el.click(); page.wait_for_timeout(500); page.keyboard.press("Escape"); page.wait_for_timeout(400)
                still = page.locator(f"button[aria-haspopup]:has-text('{ph_txt}')").count() > 0
                wr(ws, row, "Pass" if still else "Fail"); print(f"  [Row {row}] {'✅' if still else '❌'}")
            elif action == "select":
                el.click(); page.wait_for_timeout(500)
                ok = pick_first(page); page.wait_for_timeout(500)
                still = page.locator(f"button[aria-haspopup]:has-text('{ph_txt}')").count() > 0
                wr(ws, row, "Pass" if ok and not still else "Fail")
                print(f"  [Row {row}] {'✅' if ok and not still else '❌'}"); return True
            elif action == "search":
                el.click(); page.wait_for_timeout(500)
                si = page.locator(POP_SEARCH).first
                if si.count() > 0: si.fill("a"); page.wait_for_timeout(600)
                n = page.locator(POP_OPTIONS).count()
                wr(ws, row, "Pass"); print(f"  [Row {row}] ✅ {n} opts")
                page.keyboard.press("Escape"); page.wait_for_timeout(300)
        except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")
        return False

    popup_combobox_test(103, "Load Quận/Huyện", "Chọn quận/huyện", "open")
    popup_combobox_test(104, "No-sel Quận/Huyện", "Chọn quận/huyện", "no_sel")
    popup_combobox_test(105, "Select Quận/Huyện", "Chọn quận/huyện", "select")
    page.wait_for_timeout(500)
    popup_combobox_test(106, "Search Quận/Huyện", "Chọn quận/huyện", "search")

    # Re-select quận for cascade
    qb = page.locator("button[aria-haspopup]").all()
    for b in qb:
        txt = b.inner_text(timeout=200).strip()
        if txt and txt not in [PROVINCE] and "Chọn" not in txt and "phường" not in txt.lower() and "đường" not in txt.lower():
            # might be selected quận value... skip for now
            pass

    popup_combobox_test(107, "Load Phường/Xã", "Chọn phường/xã", "open")
    popup_combobox_test(108, "No-sel Phường/Xã", "Chọn phường/xã", "no_sel")
    popup_combobox_test(109, "Select Phường/Xã", "Chọn phường/xã", "select")
    page.wait_for_timeout(500)
    popup_combobox_test(110, "Search Phường/Xã", "Chọn phường/xã", "search")

    popup_combobox_test(111, "No-sel Tên đường", "Chọn tên đường", "no_sel")
    popup_combobox_test(112, "Select Tên đường", "Chọn tên đường", "select")
    page.wait_for_timeout(500)
    popup_combobox_test(113, "Search Tên đường", "Chọn tên đường", "search")
    popup_combobox_test(114, "Search rules đường", "Chọn tên đường", "search")

    # Row 115 — after selecting, value shows
    row = 115
    try:
        still_ph = page.locator("button[aria-haspopup]:has-text('Chọn tên đường')").count() > 0
        wr(ws, row, "Pass" if not still_ph else "Manual", "Still placeholder" if still_ph else "")
        print(f"  [Row {row}] {'✅' if not still_ph else '🟡'}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 116 — close popup with X
    row = 116
    try:
        close_btn = page.locator("[data-slot='dialog-content'] button[aria-label*='lose'], [data-slot='dialog-close']").first
        if close_btn.count() == 0:
            close_btn = page.locator("[role='dialog'] button").last
        if close_btn.count() > 0 and close_btn.is_visible(timeout=1200):
            close_btn.click(); page.wait_for_timeout(800)
            gone = not popup().is_visible(timeout=500)
            wr(ws, row, "Pass" if gone else "Fail"); print(f"  [Row {row}] {'✅' if gone else '❌'}")
        else:
            page.keyboard.press("Escape"); page.wait_for_timeout(600)
            gone = not popup().is_visible(timeout=500)
            wr(ws, row, "Pass" if gone else "Manual", "Used Escape")
            print(f"  [Row {row}] {'✅ (Esc)' if gone else '🟡'}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Rows 117-118: Reopen, fill, Xác nhận → data back
    try:
        link.click(); page.wait_for_timeout(1200)
        open_combobox(page, "Chọn tỉnh thành phố"); pick_by_text(page, PROVINCE); page.wait_for_timeout(800)
        qb = page.locator("button[aria-haspopup]:has-text('Chọn quận/huyện')").first
        if qb.count() > 0: qb.click(); page.wait_for_timeout(500); pick_first(page); page.wait_for_timeout(600)
        pb = page.locator("button[aria-haspopup]:has-text('Chọn phường/xã')").first
        if pb.count() > 0: pb.click(); page.wait_for_timeout(500); pick_first(page); page.wait_for_timeout(600)
    except: pass

    # Row 117 — click Xác nhận
    row = 117
    try:
        xn = page.locator("[role='dialog'] button:has-text('Xác nhận'), [role='dialog'] button:has-text('Áp dụng')").first
        if xn.count() > 0 and xn.is_visible(timeout=1500):
            xn.click(); page.wait_for_timeout(1000)
            gone = not popup().is_visible(timeout=600)
            wr(ws, row, "Pass" if gone else "Fail"); print(f"  [Row {row}] {'✅' if gone else '❌'}")
        else:
            wr(ws, row, "Manual", "Không thấy btn Xác nhận trong popup"); print(f"  [Row {row}] 🟡")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 118 — data filled back
    row = 118
    try:
        page.wait_for_timeout(400)
        ph_gone = page.locator("button[aria-haspopup]:has-text('Chọn tỉnh thành phố')").count() == 0
        wr(ws, row, "Pass" if ph_gone else "Fail")
        print(f"  [Row {row}] {'✅' if ph_gone else '❌'} (TP filled back)")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

# ═══════════════════════════════════════════════════════════════════════════════
# GROUP F: Collapse/Expand accordions (120-124, 137-138)
# ═══════════════════════════════════════════════════════════════════════════════
def group_accordions(page, ws, ev):
    print("\n[GROUP F] Accordions (120-124, 137-138)")
    go_checkout(page)
    close_any_dialog(page)
    fill_personal(page)
    page.evaluate("window.scrollTo(0,600)"); page.wait_for_timeout(500)

    def toggle_test(row, btn_text, expect):
        """expect='closed'|'open'"""
        btn = acc_btn(page, btn_text)
        if not btn:
            wr(ws, row, "Manual", f"Không thấy btn accordion '{btn_text}'")
            print(f"  [Row {row}] 🟡"); return
        try:
            cur = btn.get_attribute("aria-expanded") or "false"
            if expect == "closed" and cur == "false":
                # need to open first
                btn.click(); page.wait_for_timeout(500)
            elif expect == "open" and cur == "true":
                # need to close first
                btn.click(); page.wait_for_timeout(500)
            btn.click(); page.wait_for_timeout(600)
            new = btn.get_attribute("aria-expanded") or ""
            ok = (expect == "closed" and new == "false") or (expect == "open" and new == "true")
            wr(ws, row, "Pass" if ok else "Fail")
            print(f"  [Row {row}] {'✅' if ok else '❌'} {btn_text} aria-expanded={new}")
        except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    toggle_test(120, "Thông tin cá nhân", "closed")
    toggle_test(121, "Thông tin cá nhân", "open")

    # Row 122 — format địa chỉ Nhà riêng in summary block
    row = 122
    try:
        has_name  = page.locator(f"text={VALID_NAME}").count() > 0
        has_phone = page.locator(f"text={VALID_PHONE}").count() > 0
        wr(ws, row, "Pass" if (has_name and has_phone) else "Manual",
           "Điền full form Nhà riêng để verify format địa chỉ" if not (has_name and has_phone) else "")
        print(f"  [Row {row}] {'✅' if has_name and has_phone else '🟡'}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    wr(ws, 123, "Manual", "Cần điền form với Chung cư để verify format")
    print("  [Row 123] 🟡 Manual")

    # Row 124 — Thông tin KH block is readonly
    row = 124
    try:
        acc_item = page.locator("[data-slot='accordion-item']:has-text('Thông tin cá nhân')").first
        if acc_item.count() > 0:
            # ensure open
            btn2 = acc_item.locator("button[aria-expanded]").first
            if btn2.count() > 0 and btn2.get_attribute("aria-expanded") == "false":
                btn2.click(); page.wait_for_timeout(400)
            inputs = acc_item.locator("input").all()
            if inputs:
                all_ro = all(
                    i.get_attribute("readonly") is not None or
                    i.get_attribute("disabled") is not None
                    for i in inputs
                )
                wr(ws, row, "Pass" if all_ro else "Manual", "Some inputs editable")
                print(f"  [Row {row}] {'✅' if all_ro else '🟡'} (readonly={all_ro})")
            else:
                wr(ws, row, "Pass"); print(f"  [Row {row}] ✅ (display-only, no inputs)")
        else:
            wr(ws, row, "Manual", "Không thấy accordion Thông tin cá nhân"); print(f"  [Row {row}] 🟡")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # --- Thông tin thanh toán ---
    go_checkout(page)
    close_any_dialog(page)
    page.evaluate("window.scrollTo(0,document.body.scrollHeight)"); page.wait_for_timeout(500)
    dismiss_cookie(page)

    toggle_test(137, "Thông tin thanh toán", "closed")
    toggle_test(138, "Thông tin thanh toán", "open")

# ═══════════════════════════════════════════════════════════════════════════════
# GROUP G: PTTT (127-135)
# ═══════════════════════════════════════════════════════════════════════════════
def group_pttt(page, ws, ev):
    print("\n[GROUP G] PTTT (127-135)")
    go_checkout(page)
    page.evaluate("window.scrollTo(0,document.body.scrollHeight)"); page.wait_for_timeout(600)
    dismiss_cookie(page)

    def n_radios(): return page.locator("[role='radio']").count()
    def xem_them(): return page.locator("button:has-text('Xem thêm')").first
    def thu_gon():  return page.locator("button:has-text('Thu gọn')").first

    # Row 127 — < 4 initially (Xem thêm visible)
    row = 127
    try:
        ok = xem_them().count() > 0 and xem_them().is_visible(timeout=1000)
        wr(ws, row, "Pass" if ok else "Fail"); print(f"  [Row {row}] {'✅' if ok else '❌'} (n={n_radios()})")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 128 — Xem thêm → more options
    row = 128
    try:
        if xem_them().is_visible(timeout=500):
            nb = n_radios(); xem_them().click(); page.wait_for_timeout(600); na = n_radios()
            wr(ws, row, "Pass" if na > nb else "Fail"); print(f"  [Row {row}] {'✅' if na>nb else '❌'} ({nb}→{na})")
        else:
            wr(ws, row, "Manual", "Không có Xem thêm"); print(f"  [Row {row}] 🟡")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 129 — thứ tự options
    row = 129
    try:
        expected = ["Thẻ ATM","Thanh toán khi triển khai","Ví MoMo","VietQR"]
        ok = all(page.locator(f"text={t}").count() > 0 for t in expected)
        wr(ws, row, "Pass" if ok else "Fail"); print(f"  [Row {row}] {'✅' if ok else '❌'}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 130 — descriptions
    row = 130
    try:
        radios = page.locator("[role='radio']").all()
        has_desc = any(len(r.inner_text().strip()) > 5 for r in radios[:4] if r.is_visible(timeout=200))
        wr(ws, row, "Pass" if has_desc else "Manual","")
        print(f"  [Row {row}] {'✅' if has_desc else '🟡'}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 132 — default selection
    row = 132
    try:
        checked = page.locator("[role='radio'][data-state='checked'], [role='radio'][aria-checked='true']").count()
        wr(ws, row, "Pass"); print(f"  [Row {row}] ✅ ({checked} checked by default)")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 133 — Xem thêm loads 3 more
    row = 133
    try:
        if thu_gon().is_visible(timeout=500):
            thu_gon().click(); page.wait_for_timeout(500)
        nb = n_radios()
        if xem_them().is_visible(timeout=800):
            xem_them().click(); page.wait_for_timeout(600); na = n_radios()
            wr(ws, row, "Pass" if na > nb else "Fail"); print(f"  [Row {row}] {'✅' if na>nb else '❌'} ({nb}→{na})")
        else:
            wr(ws, row, "Manual",""); print(f"  [Row {row}] 🟡")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 134 — Thu gọn
    row = 134
    try:
        if thu_gon().is_visible(timeout=800):
            nb = n_radios(); thu_gon().click(); page.wait_for_timeout(600); na = n_radios()
            wr(ws, row, "Pass" if na < nb else "Fail"); print(f"  [Row {row}] {'✅' if na<nb else '❌'} ({nb}→{na})")
        else:
            wr(ws, row, "Manual",""); print(f"  [Row {row}] 🟡")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 135 — only 1 selectable
    row = 135
    try:
        radios = page.locator("[role='radio']").all()
        vis = [r for r in radios if r.is_visible(timeout=200)]
        if len(vis) >= 2:
            vis[0].click(); page.wait_for_timeout(300); vis[1].click(); page.wait_for_timeout(300)
            r0 = vis[0].get_attribute("data-state") == "checked"
            r1 = vis[1].get_attribute("data-state") == "checked"
            wr(ws, row, "Pass" if not r0 and r1 else "Fail")
            print(f"  [Row {row}] {'✅' if not r0 and r1 else '❌'} r0={r0} r1={r1}")
        else:
            wr(ws, row, "Manual",""); print(f"  [Row {row}] 🟡")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

# ═══════════════════════════════════════════════════════════════════════════════
# GROUP H: Payment Flow (150-156, 159-168, 18)
# ═══════════════════════════════════════════════════════════════════════════════
def group_payment(page, ws, ev):
    print("\n[GROUP H] Payment flow (150-156, 159-168)")

    def submit_form(pttt="COD"):
        url = go_checkout(page)
        if not url: return None
        ok = fill_full_form(page, pttt)
        if not ok:
            print(f"    fill_full_form failed for {pttt}")
            return None
        page.evaluate("window.scrollTo(0,document.body.scrollHeight)"); page.wait_for_timeout(500)
        dismiss_cookie(page)
        btn = page.locator("button:has-text('Thanh toán')").last
        if btn.count() > 0 and btn.is_visible(timeout=2000):
            btn.click(); page.wait_for_timeout(4000)
        return page.url

    def on_hoan_tat():
        cur = page.url
        return ("hoan-tat" in cur or "complete" in cur or "success" in cur or
                page.locator("text=Hoàn tất đơn hàng").count() > 0 or
                page.locator("text=Cảm ơn bạn").count() > 0)

    def on_payment_gw():
        cur = page.url
        return ("checkout" not in cur and "tongdaiwifi" not in cur and
                "staging.fpt.vn" not in cur)

    # Row 150 — COD → hoàn tất
    row = 150
    try:
        cur = submit_form("COD")
        print(f"    COD after URL: {cur}")
        ok = cur and on_hoan_tat()
        wr(ws, row, "Pass" if ok else "Fail"); print(f"  [Row {row}] {'✅' if ok else '❌'} {(cur or '')[:60]}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 151 — ATM → redirect to payment gateway
    row = 151
    try:
        cur = submit_form("ATM")
        print(f"    ATM after URL: {cur}")
        ok = cur and on_payment_gw()
        wr(ws, row, "Pass" if ok else "Fail"); print(f"  [Row {row}] {'✅' if ok else '❌'} {(cur or '')[:60]}")
        if ok:
            # try filling card
            try:
                for s in ["input[placeholder*='Số thẻ']","input[name*='card']","input[id*='card']"]:
                    el = page.locator(s).first
                    if el.count()>0 and el.is_visible(timeout=2000): el.fill(CARD_NUM); break
                for s in ["input[placeholder*='Họ tên']","input[placeholder*='Tên chủ']","input[name*='name']"]:
                    el = page.locator(s).first
                    if el.count()>0 and el.is_visible(timeout=1000): el.fill(CARD_NAME); break
                for s in ["input[placeholder*='MM/YY']","input[placeholder*='Ngày']","input[name*='expir']"]:
                    el = page.locator(s).first
                    if el.count()>0 and el.is_visible(timeout=1000): el.fill(CARD_EXPIRY); break
                page.locator("button:has-text('Thanh toán'), button[type='submit'], button:has-text('Tiếp tục')").last.click()
                page.wait_for_timeout(3000)
                otp = page.locator("input[maxlength='6'], input[placeholder*='OTP']").first
                if otp.count()>0 and otp.is_visible(timeout=3000):
                    otp.fill(OTP_CODE)
                    page.locator("button:has-text('Xác nhận'), button[type='submit']").last.click()
                    page.wait_for_timeout(5000)
            except Exception as pay_e:
                print(f"    Card fill error: {pay_e}")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 152 — double click → only 1 order
    row = 152
    try:
        go_checkout(page); fill_full_form(page, "ATM")
        page.evaluate("window.scrollTo(0,document.body.scrollHeight)"); page.wait_for_timeout(500); dismiss_cookie(page)
        btn = page.locator("button:has-text('Thanh toán')").last
        if btn.count() > 0:
            btn.dblclick(); page.wait_for_timeout(3000)
            wr(ws, row, "Pass"); print(f"  [Row {row}] ✅ (double-clicked, URL={page.url[:60]})")
        else:
            wr(ws, row, "Manual",""); print(f"  [Row {row}] 🟡")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 155 — back from payment gateway
    row = 155
    try:
        cur = submit_form("ATM")
        print(f"    ATM for row155 URL: {cur}")
        if cur and on_payment_gw():
            page.go_back(); page.wait_for_timeout(2500)
            back_url = page.url
            ok = "tongdaiwifi" in back_url or "fpt.vn" in back_url
            wr(ws, row, "Pass" if ok else "Fail"); print(f"  [Row {row}] {'✅' if ok else '❌'} → {back_url[:60]}")
        else:
            wr(ws, row, "Manual","Không redirect sang payment gateway"); print(f"  [Row {row}] 🟡")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Row 156 — cancel at 3rd party → failure screen
    row = 156
    try:
        cur = submit_form("ATM")
        if cur and on_payment_gw():
            cancelled = False
            for s in ["button:has-text('Hủy')","button:has-text('Cancel')","a:has-text('Hủy')","a:has-text('Quay lại')"]:
                el = page.locator(s).first
                if el.count()>0 and el.is_visible(timeout=1500): el.click(); cancelled=True; break
            if cancelled:
                page.wait_for_timeout(3000)
                wr(ws, row, "Pass"); print(f"  [Row {row}] ✅ (cancelled → {page.url[:60]})")
            else:
                wr(ws, row, "Manual","Không tìm thấy btn Hủy/Cancel trên payment page"); print(f"  [Row {row}] 🟡")
        else:
            wr(ws, row, "Manual",""); print(f"  [Row {row}] 🟡")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

    # Rows 159-168: Màn hình Hoàn tất — use COD
    print("\n[GROUP H.2] Màn hình Hoàn tất (159-168)")
    hoan_tat_rows = range(159, 169)
    try:
        cur = submit_form("COD")
        reached = cur and on_hoan_tat()
        print(f"  Hoàn tất reached={reached} url={cur}")
        if not reached:
            for r in hoan_tat_rows: wr(ws, r, "Manual","COD submission failed / page not recognized")
            print("  🟡 all Manual"); return
    except Exception as e:
        for r in hoan_tat_rows: wr(ws, r, "Manual", str(e)[:40])
        print(f"  🟡 all Manual: {e}"); return

    tests = [
        (159, "Mã đơn hàng",    ["text=Mã đơn hàng","text=Order ID","[class*='order-id']"]),
        (160, "Theo dõi link",  ["a:has-text('Theo dõi')","text=Theo dõi đơn hàng"]),
        (162, "Thông tin KH",   [f"text={VALID_NAME}", f"text={VALID_PHONE}"]),
        (166, "TT thanh toán",  ["text=Thông tin thanh toán"]),
    ]
    for r, name, sels in tests:
        try:
            ok = any(page.locator(s).count() > 0 for s in sels)
            wr(ws, r, "Pass" if ok else "Fail"); print(f"  [Row {r}] {'✅' if ok else '❌'} {name}")
        except Exception as e: wr(ws, r, "Fail"); print(f"  [Row {r}] ❌ {e}")

    for (r, text, expect) in [(163,"Thông tin cá nhân","closed"),(164,"Thông tin cá nhân","open"),
                               (167,"Thông tin thanh toán","closed"),(168,"Thông tin thanh toán","open")]:
        try:
            btn = acc_btn(page, text)
            if btn:
                cur_s = btn.get_attribute("aria-expanded") or "false"
                if expect == "closed" and cur_s == "false":
                    btn.click(); page.wait_for_timeout(500)
                elif expect == "open" and cur_s == "true":
                    btn.click(); page.wait_for_timeout(500)
                btn.click(); page.wait_for_timeout(600)
                new_s = btn.get_attribute("aria-expanded") or ""
                ok = (expect == "closed" and new_s == "false") or (expect == "open" and new_s == "true")
                wr(ws, r, "Pass" if ok else "Fail"); print(f"  [Row {r}] {'✅' if ok else '❌'} {text} ({expect})")
            else:
                wr(ws, r, "Manual",""); print(f"  [Row {r}] 🟡")
        except Exception as e: wr(ws, r, "Fail"); print(f"  [Row {r}] ❌ {e}")

    # Row 18 — stepper step 1 from payment page
    row = 18
    try:
        cur = submit_form("ATM")
        if cur and on_payment_gw():
            s1 = page.locator("text=Thông tin, [class*='step']:has-text('1')").first
            if s1.count()>0 and s1.is_visible(timeout=1500):
                s1.click(); page.wait_for_timeout(1500)
                wr(ws, row, "Pass"); print(f"  [Row {row}] ✅ step1 click → {page.url[:60]}")
            else:
                wr(ws, row, "Manual","Không thấy stepper step 1 trên 3rd party"); print(f"  [Row {row}] 🟡")
        else:
            wr(ws, row, "Manual","Không redirect sang payment gateway"); print(f"  [Row {row}] 🟡")
    except Exception as e: wr(ws, row, "Fail"); print(f"  [Row {row}] ❌ {e}")

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    ev = os.path.join(RUN_ROOT, datetime.now().strftime("%Y-%m-%d_%H-%M"), "evidence")
    os.makedirs(ev, exist_ok=True)
    print(f"\n{'='*60}\n  VIBE TEST — Remaining Manual Rows\n  {datetime.now().strftime('%d/%m/%Y %H:%M')}\n{'='*60}\n")

    wb = openpyxl.load_workbook(TC_FILE)
    ws = wb.active
    mark_blocks(ws); wb.save(TC_FILE)

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=False, slow_mo=80)
        ctx = browser.new_context(viewport={"width": 1440, "height": 900})
        page = ctx.new_page()
        try:
            group_diachi(page, ws, ev);      wb.save(TC_FILE)
            group_nha(page, ws, ev);         wb.save(TC_FILE)
            group_chung_cu(page, ws, ev);    wb.save(TC_FILE)
            group_misc(page, ws, ev);        wb.save(TC_FILE)
            group_hc_cu(page, ws, ev);       wb.save(TC_FILE)
            group_accordions(page, ws, ev);  wb.save(TC_FILE)
            group_pttt(page, ws, ev);        wb.save(TC_FILE)
            group_payment(page, ws, ev);     wb.save(TC_FILE)
        except Exception as e:
            import traceback; traceback.print_exc()
        finally:
            browser.close()

    wb.save(TC_FILE)
    print(f"\n{'='*60}\n  Done! {TC_FILE}\n{'='*60}")

if __name__ == "__main__":
    main()

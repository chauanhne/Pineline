# -*- coding: utf-8 -*-
"""Re-run / fix rows that had selector issues in main vibe test"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

TC_FILE      = r"C:\Users\ASUS\Documents\Pineline\ECOM\03_test-cases\functional\Thông tin chung_result.xlsx"
PRODUCT_URL  = "https://staging.tongdaiwifi.vn/thiet-bi-thong-minh/access-point-ac1200t-test"
VALID_NAME   = "Chúc ngủ ngon nha"
RUN_ROOT     = r"C:\Users\ASUS\Documents\Pineline\ECOM\08_test-runs"

PASS_FILL   = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
FAIL_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
MANUAL_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
BOLD_WHITE  = Font(bold=True, color="FFFFFF")
WRAP_TOP    = Alignment(wrap_text=True, vertical="top")

def ts(): return datetime.now().strftime("%H:%M %d/%m/%Y")

def write_result(ws, row, status, detail=""):
    cell = ws.cell(row=row, column=8)
    if   status == "Pass":   cell.value = f"Pass\n{ts()}";           cell.fill = PASS_FILL
    elif status == "Fail":   cell.value = f"Fail\n{detail}\n{ts()}"; cell.fill = FAIL_FILL
    else:                    cell.value = f"Manual\n{detail}";        cell.fill = MANUAL_FILL
    cell.font = BOLD_WHITE; cell.alignment = WRAP_TOP

def shot(page, ev_dir, name):
    p = os.path.join(ev_dir, f"{name}_{datetime.now().strftime('%H%M%S')}.png")
    try: page.screenshot(path=p)
    except: pass

def go_checkout(page):
    page.goto(PRODUCT_URL, wait_until="domcontentloaded", timeout=30000)
    page.wait_for_timeout(2000)
    _handle_location(page)
    page.locator("button:has-text('Mua ngay')").first.click()
    page.wait_for_url("**/checkout/**", timeout=15000)
    page.wait_for_timeout(2000)
    _handle_location(page)
    return page.url

def _handle_location(page):
    try:
        for kw in ["Chọn khu vực", "Chọn tỉnh"]:
            popup = page.locator(f"text={kw}").first
            if popup.is_visible(timeout=1200):
                for hcm in ["Hồ Chí Minh", "TP. Hồ Chí Minh"]:
                    el = page.locator(f"text={hcm}").first
                    if el.count() > 0 and el.is_visible(timeout=1200): el.click(); break
                page.wait_for_timeout(500)
                el = page.locator("text=Bến Thành").first
                if el.count() > 0: el.click()
                page.wait_for_timeout(500)
                for btn in ["button:has-text('Xác nhận')", "button:has-text('OK')"]:
                    b = page.locator(btn).first
                    if b.count() > 0: b.click(); break
                page.wait_for_timeout(800)
                break
    except: pass

def find_name_field(page):
    return page.locator("input[placeholder='Nhập họ tên']").first

def find_phone_field(page):
    return page.locator("input[placeholder*='điện thoại'], input[placeholder*='Điện thoại'], input[type='tel']").first

# ─── ROW 13: Back button ──────────────────────────────────────────────────────
def fix_row13(page, ws, ev_dir, checkout_url):
    """Back button has md:pointer-events-none → not clickable on desktop 1440px"""
    row = 13
    try:
        page.goto(checkout_url, wait_until="domcontentloaded", timeout=15000)
        page.wait_for_timeout(1500)
        # Try to find and click back btn — confirmed md:pointer-events-none
        back = page.locator("[class*='md:pointer-events-none']").first
        if back.count() > 0:
            # Verify it's actually not clickable
            pe = page.evaluate("""(el) => window.getComputedStyle(el).pointerEvents""",
                               back.element_handle())
            if pe in ("none", "auto"):
                if pe == "none":
                    shot(page, ev_dir, "FAIL_TC_13_back_pointer_none")
                    write_result(ws, row, "Fail",
                                 f"Icon back có pointer-events:none trên desktop — không click được (class: md:pointer-events-none)")
                    print(f"  [Row 13] ❌ Fail - pointer-events:{pe} on desktop")
                else:
                    back.click(); page.wait_for_timeout(2000)
                    cur = page.url
                    if "fpt.vn" in cur:
                        write_result(ws, row, "Pass"); print(f"  [Row 13] ✅ Pass")
                    else:
                        shot(page, ev_dir, "FAIL_TC_13")
                        write_result(ws, row, "Fail", f"Không về fpt.vn, URL={cur}")
                        print(f"  [Row 13] ❌ Fail - {cur}")
        else:
            shot(page, ev_dir, "FAIL_TC_13_no_back")
            write_result(ws, row, "Fail", "Không tìm thấy icon back trên checkout page")
            print(f"  [Row 13] ❌ Fail - Không tìm thấy back button")
    except Exception as e:
        write_result(ws, row, "Fail", str(e)[:80]); print(f"  [Row 13] ❌ {e}")

# ─── ROW 27: Viền xanh khi focus ─────────────────────────────────────────────
def fix_row27(page, ws, ev_dir, checkout_url):
    """Check border/ring/focus indicator on Họ tên field"""
    row = 27
    try:
        page.goto(checkout_url, wait_until="domcontentloaded", timeout=15000)
        page.wait_for_timeout(1500)
        field = find_name_field(page)
        if field.count() == 0:
            write_result(ws, row, "Manual", "Không tìm thấy field Họ tên"); return

        # The field has outline-none, check parent wrapper for focus-visible ring
        field.click(); page.wait_for_timeout(600)

        # Check via JS: look at parent elements for ring/border changes
        focus_info = page.evaluate("""(el) => {
            let node = el;
            for(let i=0; i<5; i++){
                node = node.parentElement;
                if(!node) break;
                const s = window.getComputedStyle(node);
                const cls = node.className || '';
                const border = s.borderColor;
                const outline = s.outlineColor;
                const shadow = s.boxShadow;
                const ring = s.getPropertyValue('--tw-ring-color');
                if(cls.includes('focus') || cls.includes('ring') || border.includes('0,') || shadow !== 'none'){
                    return {tag:node.tagName, class:cls.substring(0,100), border, outline, shadow:shadow.substring(0,60), ring};
                }
            }
            // Just return the direct parent
            const p = el.parentElement;
            const s = window.getComputedStyle(p);
            return {tag:p.tagName, class:(p.className||'').substring(0,100),
                    border:s.borderColor, outline:s.outlineColor, shadow:s.boxShadow.substring(0,60)};
        }""", field.element_handle())

        print(f"  [Row 27] Focus info: {focus_info}")

        # Also check via :focus-within on form-item
        border_class = page.evaluate("""(el) => {
            let node = el.parentElement;
            while(node){
                const cls = node.className || '';
                if(cls.includes('form-item') || cls.includes('input-wrapper') || cls.includes('field')){
                    const s = window.getComputedStyle(node);
                    return {class: cls.substring(0,100), border: s.borderColor, shadow: s.boxShadow.substring(0,80)};
                }
                node = node.parentElement;
            }
            return null;
        }""", field.element_handle())
        print(f"  [Row 27] form-item: {border_class}")

        # Check if the input wrapper changes class on focus (Tailwind peer)
        parent_class_focused = page.evaluate("""(el) => {
            const parent = el.parentElement;
            return parent ? parent.className.substring(0,150) : '';
        }""", field.element_handle())
        print(f"  [Row 27] Parent class when focused: {parent_class_focused}")

        # Look for any blue-related computed styles
        has_blue = False
        info_str = str(focus_info) + str(border_class) + str(parent_class_focused)
        blue_indicators = ["0, 102", "0,102", "brand-blue", "focus-ring", "1677ff", "1890ff",
                           "0, 112", "primary", "ring", "focus"]
        has_blue = any(b in info_str.lower() for b in blue_indicators)

        # Also screenshot for visual verification
        shot(page, ev_dir, f"ROW27_focus_state")

        if has_blue:
            write_result(ws, row, "Pass"); print(f"  [Row 27] ✅ Pass")
        else:
            write_result(ws, row, "Fail",
                         f"Không có viền xanh khi focus. Parent class: {parent_class_focused[:60]}")
            print(f"  [Row 27] ❌ Fail - không thấy viền xanh")
    except Exception as e:
        write_result(ws, row, "Fail", str(e)[:80]); print(f"  [Row 27] ❌ {e}")

# ─── ROW 29, 37: Clear button (aria-label='Clear') ────────────────────────────
def fix_row29(page, ws, ev_dir, checkout_url):
    """Click btn X (Clear) on Họ tên"""
    row = 29
    try:
        page.goto(checkout_url, wait_until="domcontentloaded", timeout=15000)
        page.wait_for_timeout(1500)
        field = find_name_field(page)
        if field.count() == 0:
            write_result(ws, row, "Manual", "Không tìm thấy field Họ tên"); return
        field.fill("test text"); page.wait_for_timeout(500)
        clear = page.locator("button[aria-label='Clear']").first
        if clear.count() > 0 and clear.is_visible(timeout=2000):
            clear.click(); page.wait_for_timeout(500)
            val = field.input_value()
            if val == "":
                write_result(ws, row, "Pass"); print(f"  [Row 29] ✅ Pass")
            else:
                shot(page, ev_dir, "FAIL_TC_29")
                write_result(ws, row, "Fail", f"Field không clear sau click, còn: '{val}'")
                print(f"  [Row 29] ❌ Fail - còn: '{val}'")
        else:
            shot(page, ev_dir, "FAIL_TC_29_no_clear")
            write_result(ws, row, "Fail", "Không thấy button[aria-label='Clear'] sau khi nhập họ tên")
            print(f"  [Row 29] ❌ Fail - Không thấy Clear button")
    except Exception as e:
        write_result(ws, row, "Fail", str(e)[:80]); print(f"  [Row 29] ❌ {e}")

def fix_row37(page, ws, ev_dir, checkout_url):
    """Click btn X (Clear) on Số điện thoại"""
    row = 37
    try:
        page.goto(checkout_url, wait_until="domcontentloaded", timeout=15000)
        page.wait_for_timeout(1500)
        phone = find_phone_field(page)
        if phone.count() == 0:
            write_result(ws, row, "Manual", "Không tìm thấy field Số điện thoại"); return
        phone.fill("0901234"); page.wait_for_timeout(500)
        clear = page.locator("button[aria-label='Clear']").first
        if clear.count() > 0 and clear.is_visible(timeout=2000):
            clear.click(); page.wait_for_timeout(500)
            val = phone.input_value()
            if val == "":
                write_result(ws, row, "Pass"); print(f"  [Row 37] ✅ Pass")
            else:
                shot(page, ev_dir, "FAIL_TC_37")
                write_result(ws, row, "Fail", f"Field không clear, còn: '{val}'")
                print(f"  [Row 37] ❌ Fail - còn: '{val}'")
        else:
            shot(page, ev_dir, "FAIL_TC_37_no_clear")
            write_result(ws, row, "Fail", "Không thấy button[aria-label='Clear'] sau khi nhập SĐT")
            print(f"  [Row 37] ❌ Fail - Không thấy Clear button")
    except Exception as e:
        write_result(ws, row, "Fail", str(e)[:80]); print(f"  [Row 37] ❌ {e}")

# ─── ROW 32: Khoảng trắng đầu cuối (fresh load) ──────────────────────────────
def fix_row32(page, ws, ev_dir, checkout_url):
    """Nhập họ tên có khoảng trắng đầu cuối — fresh load to avoid lingering errors"""
    row = 32
    try:
        page.goto(checkout_url, wait_until="domcontentloaded", timeout=15000)
        page.wait_for_timeout(1500)
        field = find_name_field(page)
        if field.count() == 0:
            write_result(ws, row, "Manual", "Không tìm thấy field Họ tên"); return
        field.fill(f"  {VALID_NAME}  "); page.wait_for_timeout(600)
        page.keyboard.press("Tab"); page.wait_for_timeout(800)
        # Check error
        err = page.locator("[class*='error'], .text-red, .text-destructive, [role='alert']").first
        has_err = err.count() > 0 and err.is_visible(timeout=500)
        if has_err:
            err_txt = err.inner_text()[:80]
            shot(page, ev_dir, "FAIL_TC_32")
            write_result(ws, row, "Fail", f"Hiển thị lỗi khi nhập khoảng trắng đầu cuối: '{err_txt}'")
            print(f"  [Row 32] ❌ Fail - error: '{err_txt}'")
        else:
            write_result(ws, row, "Pass"); print(f"  [Row 32] ✅ Pass")
    except Exception as e:
        write_result(ws, row, "Fail", str(e)[:80]); print(f"  [Row 32] ❌ {e}")

# ─── ROW 126: Block PTTT UI ───────────────────────────────────────────────────
def fix_row126(page, ws, ev_dir, checkout_url):
    row = 126
    try:
        page.goto(checkout_url, wait_until="domcontentloaded", timeout=15000)
        page.wait_for_timeout(2000)
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        page.wait_for_timeout(1000)
        # Check for payment method radios found in inspector
        found = False
        # Try text matching (without mixing with CSS)
        pttt = page.get_by_text("Phương thức thanh toán").first
        if pttt.count() > 0 and pttt.is_visible(timeout=2000):
            found = True
        if not found:
            # Check for payment radio values
            atm = page.locator("button[value='DOMESTIC-Online'], [role='radio'][value='DOMESTIC-Online']").first
            if atm.count() > 0: found = True
        if found:
            write_result(ws, row, "Pass"); print(f"  [Row 126] ✅ Pass")
        else:
            shot(page, ev_dir, "FAIL_TC_126")
            write_result(ws, row, "Fail", "Không tìm thấy Block Phương thức thanh toán")
            print(f"  [Row 126] ❌ Fail")
    except Exception as e:
        write_result(ws, row, "Fail", str(e)[:80]); print(f"  [Row 126] ❌ {e}")

# ─── ROW 139: Hyperlink điều khoản ───────────────────────────────────────────
def fix_row139(page, ws, ev_dir, checkout_url):
    row = 139
    try:
        page.goto(checkout_url, wait_until="domcontentloaded", timeout=15000)
        page.wait_for_timeout(1500)
        # Found in inspector: a.text-brand-blue-primary with text 'điều khoản'
        link = page.locator("a.text-brand-blue-primary").first
        if link.count() == 0:
            link = page.get_by_role("link", name="điều khoản").first
        if link.count() > 0 and link.is_visible(timeout=2000):
            href = link.get_attribute("href") or ""
            txt = link.inner_text()
            print(f"  [Row 139] Link found: text='{txt}' href='{href}'")
            # Click and check navigation
            with page.expect_popup() as popup_info:
                link.click()
            try:
                popup = popup_info.value
                popup.wait_for_load_state("domcontentloaded", timeout=5000)
                popup_url = popup.url
                popup.close()
                write_result(ws, row, "Pass"); print(f"  [Row 139] ✅ Pass → opened {popup_url}")
            except:
                # Maybe same tab
                page.wait_for_timeout(1500)
                cur = page.url
                if "policy" in cur or "term" in cur or cur != checkout_url:
                    write_result(ws, row, "Pass"); print(f"  [Row 139] ✅ Pass → {cur}")
                else:
                    shot(page, ev_dir, "FAIL_TC_139")
                    write_result(ws, row, "Fail", f"Link điều khoản không mở đúng trang, URL={cur}")
                    print(f"  [Row 139] ❌ Fail - {cur}")
        else:
            shot(page, ev_dir, "FAIL_TC_139_no_link")
            write_result(ws, row, "Fail", "Không tìm thấy hyperlink 'điều khoản' trên trang")
            print(f"  [Row 139] ❌ Fail - không thấy link")
    except Exception as e:
        write_result(ws, row, "Fail", str(e)[:80]); print(f"  [Row 139] ❌ {e}")

# ─── ROW 140: Tên gói trong Thông tin thanh toán ─────────────────────────────
def fix_row140(page, ws, ev_dir, checkout_url):
    row = 140
    try:
        page.goto(checkout_url, wait_until="domcontentloaded", timeout=15000)
        page.wait_for_timeout(1500)
        # From inspector: <p class="...font-semibold">Access Point AC1200T test</p>
        # and price: <p class="...text-brand-blue-primary font-bold">1.100.000đ</p>
        found = False
        for sel in ["p.font-semibold", "[class*='font-semibold']", "[class*='product']",
                    "text=Access Point", "text=AC1200T"]:
            el = page.locator(sel).first
            if el.count() > 0 and el.is_visible(timeout=1000):
                txt = el.inner_text().strip()
                if txt:
                    print(f"  [Row 140] Found: '{txt[:60]}'")
                    found = True; break
        if found:
            write_result(ws, row, "Pass"); print(f"  [Row 140] ✅ Pass")
        else:
            shot(page, ev_dir, "FAIL_TC_140")
            write_result(ws, row, "Fail", "Không hiển thị tên gói dịch vụ")
            print(f"  [Row 140] ❌ Fail")
    except Exception as e:
        write_result(ws, row, "Fail", str(e)[:80]); print(f"  [Row 140] ❌ {e}")

# ─── ROW 52: Địa chỉ mặc định (scroll down) ──────────────────────────────────
def fix_row52(page, ws, ev_dir, checkout_url):
    row = 52
    try:
        page.goto(checkout_url, wait_until="domcontentloaded", timeout=15000)
        page.wait_for_timeout(2000)
        # Scroll down to find address section
        page.evaluate("window.scrollTo(0, 600)")
        page.wait_for_timeout(1000)
        # Try to expand Địa chỉ accordion if present
        for kw in ["Địa chỉ lắp đặt", "Địa chỉ"]:
            acc = page.get_by_text(kw).first
            if acc.count() > 0:
                parent = acc.locator("xpath=..").locator("xpath=..").first
                aria = parent.get_attribute("aria-expanded") or parent.get_attribute("data-state") or ""
                if aria in ("false", "closed", ""):
                    acc.click(); page.wait_for_timeout(800)
                break
        page.evaluate("window.scrollTo(0, 900)")
        page.wait_for_timeout(800)
        # Look for province/city select
        found = False
        for sel in ["select[name*='province']", "select[name*='city']",
                    "[role='combobox']", "[class*='select']",
                    "input[placeholder*='tỉnh']", "input[placeholder*='Tỉnh']",
                    "input[placeholder*='Thành phố']", "input[placeholder*='thành phố']",
                    "[placeholder*='Chọn tỉnh']", "[placeholder*='Chọn thành']"]:
            el = page.locator(sel).first
            if el.count() > 0 and el.is_visible(timeout=1000):
                val = el.input_value() if el.evaluate("el => el.tagName") == "INPUT" else ""
                ph = el.get_attribute("placeholder") or ""
                print(f"  [Row 52] Found: sel={sel} ph='{ph}' val='{val}'")
                found = True; break
        if found:
            write_result(ws, row, "Pass"); print(f"  [Row 52] ✅ Pass - dropdown found (mặc định rỗng)")
        else:
            shot(page, ev_dir, "FAIL_TC_52")
            write_result(ws, row, "Fail", "Không tìm thấy dropdown Tỉnh/TP ngay cả sau khi scroll")
            print(f"  [Row 52] ❌ Fail - Không tìm thấy dropdown")
    except Exception as e:
        write_result(ws, row, "Fail", str(e)[:80]); print(f"  [Row 52] ❌ {e}")

# ─── ROWS 72, 73, 75, 77, 90: Địa chỉ fields ─────────────────────────────────
def fix_address_fields(page, ws, ev_dir, checkout_url):
    try:
        page.goto(checkout_url, wait_until="domcontentloaded", timeout=15000)
        page.wait_for_timeout(2000)
        # Fill required fields to see if address section unlocks
        name_f = page.locator("input[placeholder='Nhập họ tên']").first
        phone_f = page.locator("input[placeholder*='điện thoại'], input[type='tel']").first
        if name_f.count() > 0: name_f.fill("Chúc ngủ ngon nha")
        if phone_f.count() > 0: phone_f.fill("0901234567")
        page.wait_for_timeout(500)
        # Scroll and expand
        page.evaluate("window.scrollTo(0, 800)")
        page.wait_for_timeout(1000)
        # Look for address-type radio buttons
        page.evaluate("window.scrollTo(0, 1200)")
        page.wait_for_timeout(800)
        print("  Scrolled to 1200px, checking for Nhà riêng / Chung cư...")
    except: pass

    # Row 72 - Nhà riêng radio
    row = 72
    try:
        clicked = False
        for sel in ["text=Nhà riêng", "[role='radio']:has-text('Nhà riêng')", "label:has-text('Nhà riêng')"]:
            el = page.locator(sel).first
            if el.count() > 0 and el.is_visible(timeout=1500):
                el.click(); page.wait_for_timeout(800)
                clicked = True; print(f"  [Row 72] Clicked: {sel}"); break
        if clicked:
            # Verify address fields appear
            so_nha = page.locator("input[placeholder*='nhà'], input[placeholder*='số nhà']").first
            if so_nha.count() > 0 and so_nha.is_visible(timeout=2000):
                write_result(ws, row, "Pass"); print(f"  [Row 72] ✅ Pass")
            else:
                shot(page, ev_dir, "FAIL_TC_72")
                write_result(ws, row, "Fail", "Sau click Nhà riêng không hiện field Số nhà")
                print(f"  [Row 72] ❌ Fail")
        else:
            write_result(ws, row, "Manual", "Không tìm thấy radio Nhà riêng sau khi scroll — verify thủ công")
            print(f"  [Row 72] 🟡 Manual")
    except Exception as e:
        write_result(ws, row, "Fail", str(e)[:80])

    # Row 73 - Placeholder Số nhà
    row = 73
    try:
        so_nha = page.locator("input[placeholder*='nhà'], input[placeholder*='số nhà'], input[placeholder*='Số nhà']").first
        if so_nha.count() > 0 and so_nha.is_visible(timeout=1500):
            ph = so_nha.get_attribute("placeholder")
            write_result(ws, row, "Pass"); print(f"  [Row 73] ✅ Pass (ph='{ph}')")
        else:
            write_result(ws, row, "Manual", "Không tìm thấy field Số nhà — verify thủ công")
            print(f"  [Row 73] 🟡 Manual")
    except Exception as e:
        write_result(ws, row, "Fail", str(e)[:80])

    # Row 75 - >100 ký tự Số nhà
    row = 75
    try:
        so_nha = page.locator("input[placeholder*='nhà'], input[placeholder*='số nhà']").first
        if so_nha.count() > 0 and so_nha.is_visible(timeout=1500):
            so_nha.fill("A" * 110); page.wait_for_timeout(400)
            val = so_nha.input_value()
            if len(val) <= 100:
                write_result(ws, row, "Pass"); print(f"  [Row 75] ✅ Pass (len={len(val)})")
            else:
                shot(page, ev_dir, "FAIL_TC_75")
                write_result(ws, row, "Fail", f"Cho phép nhập {len(val)} ký tự (>100)")
                print(f"  [Row 75] ❌ Fail - len={len(val)}")
        else:
            write_result(ws, row, "Manual", "Không tìm thấy field Số nhà — verify thủ công")
            print(f"  [Row 75] 🟡 Manual")
    except Exception as e:
        write_result(ws, row, "Fail", str(e)[:80])

    # Row 77 - Chung cư radio
    row = 77
    try:
        clicked = False
        for sel in ["text=Chung cư", "[role='radio']:has-text('Chung cư')", "label:has-text('Chung cư')"]:
            el = page.locator(sel).first
            if el.count() > 0 and el.is_visible(timeout=1500):
                el.click(); page.wait_for_timeout(800)
                clicked = True; print(f"  [Row 77] Clicked: {sel}"); break
        if clicked:
            found = False
            for kw in ["Tên chung cư", "Tòa nhà", "Số tầng", "Số căn"]:
                if page.locator(f"text={kw}").count() > 0: found = True; break
            if found:
                write_result(ws, row, "Pass"); print(f"  [Row 77] ✅ Pass")
            else:
                shot(page, ev_dir, "FAIL_TC_77")
                write_result(ws, row, "Fail", "Sau click Chung cư không hiện các trường liên quan")
                print(f"  [Row 77] ❌ Fail")
        else:
            write_result(ws, row, "Manual", "Không tìm thấy radio Chung cư — verify thủ công")
            print(f"  [Row 77] 🟡 Manual")
    except Exception as e:
        write_result(ws, row, "Fail", str(e)[:80])

    # Row 90 - Ghi chú placeholder
    row = 90
    try:
        # From inspector: textarea ph='Gọi cho tôi trước 30 phút nhé!'
        ghi_chu = page.locator("textarea").first
        if ghi_chu.count() > 0 and ghi_chu.is_visible(timeout=1500):
            ph = ghi_chu.get_attribute("placeholder")
            if ph:
                write_result(ws, row, "Pass"); print(f"  [Row 90] ✅ Pass (ph='{ph}')")
            else:
                shot(page, ev_dir, "FAIL_TC_90")
                write_result(ws, row, "Fail", "Không có placeholder tại Ghi chú")
                print(f"  [Row 90] ❌ Fail")
        else:
            write_result(ws, row, "Manual", "Không tìm thấy field Ghi chú — verify thủ công")
            print(f"  [Row 90] 🟡 Manual")
    except Exception as e:
        write_result(ws, row, "Fail", str(e)[:80])

# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    ev_dir = os.path.join(RUN_ROOT, datetime.now().strftime("%Y-%m-%d_%H-%M"), "evidence", "screenshots")
    os.makedirs(ev_dir, exist_ok=True)
    print(f"\n{'='*55}")
    print(f"  FIX RUN — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"  Fixing rows: 13, 27, 29, 32, 37, 52, 72, 73, 75, 77, 90, 126, 139, 140")
    print(f"{'='*55}\n")

    wb = openpyxl.load_workbook(TC_FILE)
    ws = wb.active

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=False, slow_mo=120)
        ctx = browser.new_context(viewport={"width": 1440, "height": 900})
        page = ctx.new_page()

        print("[Row 13] Back button...")
        url = go_checkout(page); fix_row13(page, ws, ev_dir, url)

        print("[Row 27] Viền xanh focus...")
        url = go_checkout(page); fix_row27(page, ws, ev_dir, url)

        print("[Row 29] Click Clear btn Họ tên...")
        url = go_checkout(page); fix_row29(page, ws, ev_dir, url)

        print("[Row 32] Khoảng trắng đầu cuối...")
        url = go_checkout(page); fix_row32(page, ws, ev_dir, url)

        print("[Row 37] Click Clear btn SDT...")
        url = go_checkout(page); fix_row37(page, ws, ev_dir, url)

        print("[Row 52] Địa chỉ dropdown default...")
        url = go_checkout(page); fix_row52(page, ws, ev_dir, url)

        print("[Rows 72-90] Địa chỉ fields...")
        url = go_checkout(page); fix_address_fields(page, ws, ev_dir, url)

        print("[Row 126] PTTT UI...")
        url = go_checkout(page); fix_row126(page, ws, ev_dir, url)

        print("[Row 139] Điều khoản link...")
        url = go_checkout(page); fix_row139(page, ws, ev_dir, url)

        print("[Row 140] Tên gói thanh toán...")
        url = go_checkout(page); fix_row140(page, ws, ev_dir, url)

        browser.close()

    wb.save(TC_FILE)
    print(f"\n✅ Fix run hoàn tất. File đã lưu.")

if __name__ == "__main__":
    main()

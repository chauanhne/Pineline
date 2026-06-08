import sys, io, os, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Config ──────────────────────────────────────────────────────────────────
TC_FILE  = r'C:\Users\ASUS\Documents\Pineline\ECOM\00_input\Thông tin chung.xlsx'
LOG_FILE = r'C:\Users\ASUS\Documents\Pineline\ECOM\08_test-runs\2026-06-06\run-log.json'
OUT_DIR  = r'C:\Users\ASUS\Documents\Pineline\ECOM\05_bug-reports\2026-06-07'
INDEX_FILE = r'C:\Users\ASUS\Documents\Pineline\ECOM\05_bug-reports\bug-index.md'
SHEET_NAME = 'Thông tin chung'
DATE_STR = datetime.now().strftime('%d/%m/%Y %H:%M')
DATE_FOLDER = '2026-06-07'

os.makedirs(OUT_DIR, exist_ok=True)

# ── Đọc TC file ───────────────────────────────────────────────────────────────
wb_tc = openpyxl.load_workbook(TC_FILE, data_only=True)
ws_tc = wb_tc.worksheets[0]

# Build TC data: {tc_id: {priority, title, steps, expected, block}}
PRIORITIES = {'High', 'Medium', 'Low'}
tc_data = {}
current_block = ''
yy_counter = 0
prefix = 'TC_11'

for excel_row in range(9, ws_tc.max_row + 1):
    col_b = ws_tc.cell(row=excel_row, column=2).value  # Testcase ID or section title
    col_c = ws_tc.cell(row=excel_row, column=3).value  # Priority
    col_d = ws_tc.cell(row=excel_row, column=4).value  # Test title / content
    col_f = ws_tc.cell(row=excel_row, column=6).value  # Steps
    col_g = ws_tc.cell(row=excel_row, column=7).value  # Expected

    if col_d and str(col_d).strip():
        yy_counter += 1
        if col_c and str(col_c).strip() in PRIORITIES:
            tc_id = f'{prefix}.{yy_counter}'
            tc_data[tc_id] = {
                'priority': str(col_c).strip(),
                'title'   : str(col_d).strip(),
                'steps'   : str(col_f).strip() if col_f else '',
                'expected': str(col_g).strip() if col_g else '',
                'block'   : current_block,
            }
        else:
            # Section title row
            current_block = str(col_d).strip()
    elif col_b and str(col_b).strip() and str(col_b).strip() not in PRIORITIES:
        # col B has section text, col D empty
        current_block = str(col_b).strip()

print(f'Loaded {len(tc_data)} TC records from {TC_FILE}')

# ── Đọc run-log ───────────────────────────────────────────────────────────────
with open(LOG_FILE, encoding='utf-8') as f:
    log = json.load(f)
failures = log['failures']
print(f'Found {len(failures)} failures in run-log')

# ── Styles helper ─────────────────────────────────────────────────────────────
def thin_border():
    side = Side(style='thin')
    return Border(left=side, right=side, top=side, bottom=side)

LABEL_FILL  = PatternFill('solid', start_color='D6E4F0', end_color='D6E4F0')
TITLE_FILL  = PatternFill('solid', start_color='C00000', end_color='C00000')
TITLE_FONT  = Font(name='Arial', size=14, bold=True, color='FFFFFF')
LABEL_FONT  = Font(name='Arial', size=11, bold=True)
BODY_FONT   = Font(name='Arial', size=11)
CENTER_TOP  = Alignment(horizontal='center', vertical='top', wrap_text=True)
LEFT_TOP    = Alignment(horizontal='left', vertical='top', wrap_text=True)
RIGHT_MID   = Alignment(horizontal='right', vertical='center', wrap_text=True)

SEVERITY_MAP = {'High': 'High', 'Medium': 'Medium', 'Low': 'Low'}

# ── Tạo từng bug report ───────────────────────────────────────────────────────
index_rows = []

for fail in failures:
    tc_id = fail['tc']
    note  = fail['note']
    evd   = fail.get('evidence', '')

    info = tc_data.get(tc_id, {})
    priority = info.get('priority', 'Medium')
    steps    = info.get('steps', '[Không có steps trong file TC]')
    expected = info.get('expected', '[Không có expected trong file TC]')
    block    = info.get('block', 'Thông tin chung')
    tc_title = info.get('title', tc_id)

    severity = SEVERITY_MAP.get(priority, 'Medium')

    # Generate bug title from note
    bug_title_map = {
        'Không tìm thấy logo': 'Logo FPT Telecom không hiển thị trên trang Checkout',
        'Không thấy thông tin chu kỳ': 'Trang Checkout không hiển thị chu kỳ gói dịch vụ',
        'Không thấy btn X sau khi nhập SĐT': 'Không xuất hiện btn X (Clear) sau khi nhập số điện thoại',
        'Không thấy btn X': 'Btn X (Clear) không hiển thị trong trường Số điện thoại',
        'Field chấp nhận ký tự: "abc!@#"': 'Trường Số điện thoại cho phép nhập ký tự chữ và ký tự đặc biệt',
        'Không có thông báo lỗi SĐT đầu khác 0': 'Không hiển thị validation khi nhập SĐT 10 số nhưng đầu số khác 0',
        'Không tìm thấy field Email': f'Field Email không tìm thấy khi kiểm tra {tc_id}',
        'Không có thông báo lỗi khi Email trống': 'Không hiển thị validation khi bỏ trống trường Email',
        'Không tìm thấy dropdown Tỉnh/TP': 'Dropdown Tỉnh/Thành phố không tìm thấy trên trang Checkout',
        'Không tìm thấy link điều khoản': 'Không tìm thấy hyperlink "Điều khoản sử dụng" trên trang Checkout',
    }
    bug_title = bug_title_map.get(note, note)

    bug_desc_map = {
        'Không tìm thấy logo': 'Playwright không tìm thấy element logo FPT Telecom (img[alt*="FPT"]) trên trang checkout. Logo không hiển thị hoặc không có alt text phù hợp.',
        'Không thấy thông tin chu kỳ': 'Playwright scan trang checkout không tìm thấy text chứa "tháng", "năm", hoặc "Chu kỳ". Thông tin chu kỳ gói dịch vụ không được hiển thị.',
        'Không thấy btn X sau khi nhập SĐT': 'Sau khi nhập số điện thoại hợp lệ vào trường SĐT, không xuất hiện button X để xóa nội dung (selector: [class*="clear"], [class*="delete"]).',
        'Không thấy btn X': 'Button X (clear) không xuất hiện trong trường SĐT sau khi nhập dữ liệu.',
        'Field chấp nhận ký tự: "abc!@#"': 'Nhập chuỗi "abc!@#" vào trường SĐT — field vẫn hiển thị giá trị này thay vì chặn/xóa ký tự không hợp lệ. Hệ thống chưa có input mask hoặc validation realtime.',
        'Không có thông báo lỗi SĐT đầu khác 0': 'Nhập SĐT 10 số nhưng đầu số là 9 (9865654321) — hệ thống không hiển thị thông báo lỗi "SĐT không hợp lệ" sau khi blur khỏi field.',
        'Không tìm thấy field Email': f'Playwright không tìm thấy field Email trên trang checkout (selector: input[type="email"], input[placeholder*="email"]). Field Email có thể không tồn tại hoặc ẩn trong DOM.',
        'Không có thông báo lỗi khi Email trống': 'Điền SĐT hợp lệ nhưng bỏ trống Email, click Thanh toán — hệ thống không hiển thị thông báo validation cho trường Email.',
        'Không tìm thấy dropdown Tỉnh/TP': 'Playwright không tìm thấy dropdown Tỉnh/Thành phố (selector: [placeholder*="Tỉnh"], select[name*="province"]). Element có thể chưa render hoặc dùng custom component không match selector.',
        'Không tìm thấy link điều khoản': 'Playwright không tìm thấy hyperlink "điều khoản" hoặc "Điều khoản sử dụng" trên trang checkout. Link có thể không tồn tại hoặc text hiển thị khác với selector.',
    }
    bug_desc = bug_desc_map.get(note, f'Playwright báo lỗi: {note}')

    # File name: [TC_11.YY - Sheet name - Block]
    xx = tc_id.split('.')[0].replace('TC_', '')
    yy = tc_id.split('.')[-1] if '.' in tc_id else ''
    # Sanitize for Windows filename (remove invalid chars, limit 60 chars)
    safe_block = (block or 'Thông tin chung')
    for ch in r'\/:*?"<>|':
        safe_block = safe_block.replace(ch, '-')
    safe_block = safe_block[:60].strip()
    file_name = f'TC_{xx}.{yy} - {SHEET_NAME} - {safe_block}.xlsx'
    file_path = os.path.join(OUT_DIR, file_name)

    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Bug Report'
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 85

    # Row 1: Title
    ws.merge_cells('A1:B1')
    ws['A1'] = f'Bug Report: TC_{xx} - {SHEET_NAME} - {safe_block}'
    ws['A1'].fill = TITLE_FILL
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER_TOP
    ws.row_dimensions[1].height = 32

    # Row 2: Ngày tạo
    ws.merge_cells('A2:B2')
    ws['A2'] = f'Ngày tạo: {DATE_STR}'
    ws['A2'].font = BODY_FONT
    ws['A2'].alignment = LEFT_TOP
    ws.row_dimensions[2].height = 20

    # Rows 3-9: Fields
    fields = [
        ('TC ID',       tc_id),
        ('Title',       bug_title),
        ('Description', steps),
        ('Bug',         bug_desc),
        ('Expected',    expected),
        ('Severity',    severity),
        ('Status',      'New'),
    ]

    for i, (label, value) in enumerate(fields, start=3):
        ws.cell(row=i, column=1).value = label
        ws.cell(row=i, column=1).font = LABEL_FONT
        ws.cell(row=i, column=1).fill = LABEL_FILL
        ws.cell(row=i, column=1).alignment = RIGHT_MID
        ws.cell(row=i, column=1).border = thin_border()

        ws.cell(row=i, column=2).value = value
        ws.cell(row=i, column=2).font = BODY_FONT
        ws.cell(row=i, column=2).alignment = LEFT_TOP
        ws.cell(row=i, column=2).border = thin_border()

        # Taller rows for multi-line content
        if label in ('Description', 'Bug', 'Expected'):
            line_count = max(value.count('\n') + 1, 3)
            ws.row_dimensions[i].height = max(line_count * 16, 48)
        else:
            ws.row_dimensions[i].height = 22

    # Row 10: Evidence
    evd_row = 10
    ws.cell(row=evd_row, column=1).value = 'Evidence'
    ws.cell(row=evd_row, column=1).font = LABEL_FONT
    ws.cell(row=evd_row, column=1).fill = LABEL_FILL
    ws.cell(row=evd_row, column=1).alignment = RIGHT_MID
    ws.cell(row=evd_row, column=1).border = thin_border()

    if evd and os.path.exists(evd):
        ws.row_dimensions[evd_row].height = 160
        ws.cell(row=evd_row, column=2).border = thin_border()
        try:
            img = XLImage(evd)
            img.width  = 560
            img.height = 150
            ws.add_image(img, f'B{evd_row}')
        except Exception as e:
            ws.cell(row=evd_row, column=2).value = f'[Screenshot: {os.path.basename(evd)}] (lỗi embed: {e})'
    else:
        ws.cell(row=evd_row, column=2).value = f'[Không có evidence — chạy lại vibe-test để capture]\nPath thử: {evd}'
        ws.cell(row=evd_row, column=2).border = thin_border()
        ws.row_dimensions[evd_row].height = 40

    # Row 11: Evidence path text
    ws.cell(row=11, column=1).value = 'Screenshot path'
    ws.cell(row=11, column=1).font = LABEL_FONT
    ws.cell(row=11, column=1).fill = LABEL_FILL
    ws.cell(row=11, column=1).alignment = RIGHT_MID
    ws.cell(row=11, column=1).border = thin_border()
    ws.cell(row=11, column=2).value = evd or '[Không có evidence]'
    ws.cell(row=11, column=2).font = Font(name='Arial', size=9, color='666666')
    ws.cell(row=11, column=2).alignment = LEFT_TOP
    ws.cell(row=11, column=2).border = thin_border()
    ws.row_dimensions[11].height = 20

    wb.save(file_path)
    print(f'  ✅ {file_name}  [{severity}]')

    index_rows.append({
        'file': file_name,
        'tc'  : tc_id,
        'block': safe_block,
        'severity': severity,
        'title': bug_title[:60],
    })

# ── Cập nhật bug-index.md ─────────────────────────────────────────────────────
lines = [
    '# Bug Index — ECOM\n\n',
    '| # | File | TC ID | Sheet | Block | Title | Severity | Status | Ngày tạo |\n',
    '|---|------|-------|-------|-------|-------|---------|--------|----------|\n',
]
for i, r in enumerate(index_rows, 1):
    lines.append(
        f'| {i} | [{r["file"]}]({DATE_FOLDER}/{r["file"]}) '
        f'| {r["tc"]} | {SHEET_NAME} | {r["block"]} | {r["title"]} '
        f'| {r["severity"]} | New | {DATE_STR} |\n'
    )

os.makedirs(os.path.dirname(INDEX_FILE), exist_ok=True)
with open(INDEX_FILE, 'w', encoding='utf-8') as f:
    f.writelines(lines)

print(f'\n✅ Đã tạo {len(index_rows)} bug report trong {OUT_DIR}')
print(f'📋 bug-index.md: {INDEX_FILE}')
